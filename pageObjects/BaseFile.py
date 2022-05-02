import json
import time
from faker import Faker
from datetime import date
from gzip import decompress
from openpyxl import load_workbook
from Locators.Locators import Locators
from selenium.webdriver.common.by import By
from Locators.MasterLocators import MasterLocators
from Locators.PatientLocators import PatientLocators
from Locators.PracticeLocators import PracticeLocators
from Locators.DispenserLocators import DispenserLocators
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

FilePath = "TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']
loginData = datafile["Login Credentials"]
scriptData = datafile["Script Data"]

faker = Faker()
today = date.today()
currentDate = today.strftime("%m/%d/%Y")
FirstName = faker.first_name()
LastName = faker.last_name()
PhoneNumber = faker.phone_number()
EmailAddress = FirstName + "." + LastName + "@mailinator.com"


class CommanFlow:
    def __init__(self, driver):
        self.driver = driver

    def CreatePatient(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.addPersonButton))).click()
        self.driver.find_element_by_name(PracticeLocators.firstName).send_keys(FirstName)
        self.driver.find_element_by_name(PracticeLocators.lastName).send_keys(LastName)
        self.driver.find_element_by_name(PracticeLocators.birthDate).send_keys(currentDate)
        self.driver.find_element_by_name(PracticeLocators.phoneNumber).send_keys(PhoneNumber)
        self.driver.find_element_by_name(Locators.Email).send_keys(EmailAddress.lower())
        self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
        time.sleep(5)
        if "Congratulations, you have successfully created a new Patient Account." in self.driver.page_source:
            self.driver.find_element_by_xpath(PracticeLocators.closeButton).click()
            testData.cell(2, 1).value = FirstName + " " + LastName
            testData.cell(2, 2).value = FirstName + " " + LastName + " (" + currentDate + ")"
            loginData.cell(8, 3).value = EmailAddress.lower()
            datafile.save(FilePath)
            print("\nPatient Created Successfully")
        else:
            print("\nPatient not Created")

    def CreateUser(self):
        self.driver.find_element_by_class_name(PracticeLocators.accountLinkLabel).click()
        self.driver.find_element_by_link_text(PracticeLocators.manageUser).click()
        self.driver.find_element_by_xpath(PracticeLocators.addUserButton).click()
        self.driver.find_element_by_name(PracticeLocators.firstName).send_keys(FirstName)
        self.driver.find_element_by_name(PracticeLocators.lastName).send_keys(LastName)
        self.driver.find_element_by_name(Locators.Email).send_keys(EmailAddress.lower())
        self.driver.find_element_by_name(PracticeLocators.phoneNumber).send_keys(PhoneNumber)
        self.driver.find_element_by_class_name(PracticeLocators.selectAccountType).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.clerkOption).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
        time.sleep(4)
        if "Congratulations. You have successfully created new user profile." in self.driver.page_source:
            self.driver.find_element_by_xpath(PracticeLocators.dismissButton).click()
            print("\nUser created successfully")
        else:
            print("\nUser not created")

    def PatientMainSearch(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.mainPatientSearch))).click()
        self.driver.find_element_by_xpath(PracticeLocators.mainSearchTextbox).send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_class_name(PracticeLocators.mainSearchButton).click()
        try:
            self.driver.find_element_by_xpath(PracticeLocators.tableRow).click()
            try:
                self.driver.find_element_by_link_text(PracticeLocators.viewDetails).click()
            except:
                print("\nOrder not created for " + testData.cell(2, 1).value + ".")
        except:
            print("\nPatient: " + testData.cell(2, 1).value + " not found...")

    def ChangeProductPrice(self):
        self.driver.find_element_by_class_name(PracticeLocators.accountLinkLabel).click()
        self.driver.find_element_by_link_text(DispenserLocators.productList).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.tableRow).click()
        self.driver.find_element_by_xpath(DispenserLocators.updateButton).click()
        self.driver.find_element_by_xpath(DispenserLocators.changePriceButton).click()
        time.sleep(3)
        if "You have successfully updated the price" in self.driver.page_source:
            self.driver.find_element_by_xpath(PracticeLocators.closeButton).click()
            print('\nProduct price changed successfully')
        else:
            print('\nProduct price not changed')

    def OnetimeRXSkipPayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.CreateOnetimeRXButton).click()
        self.driver.find_element_by_name(PracticeLocators.OnetimeSearchMedicine).send_keys(scriptData.cell(2, 1).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 1).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.skipPayment).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def OnetimeOTCSkipPayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.CreateOnetimeRXButton).click()
        self.driver.find_element_by_name(PracticeLocators.OnetimeSearchMedicine).send_keys(scriptData.cell(2, 2).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 2).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.skipPayment).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def OnetimeCompoundSkipPayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.CreateOnetimeRXButton).click()
        self.driver.find_element_by_name(PracticeLocators.OnetimeSearchMedicine).send_keys(scriptData.cell(2, 3).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 3).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.skipPayment).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def SubscriptionRXSkipPayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        el = self.driver.find_element_by_xpath(PracticeLocators.createSubscriptionRxButton)
        self.driver.execute_script('arguments[0].click()', el)
        self.driver.find_element_by_xpath(PracticeLocators.subscriptionProductSearch).send_keys(
            scriptData.cell(2, 1).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 1).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.skipPayment).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(DispenserLocators.selectServiceType).click()
        time.sleep(1)
        self.driver.find_element_by_css_selector(DispenserLocators.pickUpPerson).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def SubscriptionOTCSkipPayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        el = self.driver.find_element_by_xpath(PracticeLocators.createSubscriptionRxButton)
        self.driver.execute_script('arguments[0].click()', el)
        self.driver.find_element_by_xpath(PracticeLocators.subscriptionProductSearch).send_keys(
            scriptData.cell(2, 2).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 2).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.skipPayment).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def SubscriptionCompoundSkipPayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        el = self.driver.find_element_by_xpath(PracticeLocators.createSubscriptionRxButton)
        self.driver.execute_script('arguments[0].click()', el)
        self.driver.find_element_by_xpath(PracticeLocators.subscriptionProductSearch).send_keys(
            scriptData.cell(2, 3).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 3).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        ExpiryDate = self.driver.find_elements(By.XPATH, PracticeLocators.surgeryDate)
        if ExpiryDate:
            ExpiryDate[0].send_keys(currentDate)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.skipPayment).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def OnetimeRXProvidePayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.CreateOnetimeRXButton).click()
        self.driver.find_element_by_name(PracticeLocators.OnetimeSearchMedicine).send_keys(scriptData.cell(2, 1).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 1).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.providePayment).click()
        EditPayment = self.driver.find_elements(By.XPATH, DispenserLocators.editPaymentButton)
        if EditPayment:
            EditPayment[0].click()
        cardName = self.driver.find_element_by_name(PracticeLocators.cardName)
        cardName.clear()
        cardName.send_keys(testData.cell(2, 7).value)
        cardNumber = self.driver.find_element_by_name(PracticeLocators.maskCardNumberField)
        cardNumber.clear()
        cardNumber.send_keys(testData.cell(2, 8).value)
        cardCVV = self.driver.find_element_by_name(PracticeLocators.cardCVV)
        cardCVV.clear()
        cardCVV.send_keys(testData.cell(2, 9).value)

        Address1 = self.driver.find_element_by_name(PracticeLocators.addressLine1)
        Address1.clear()
        Address1.send_keys(testData.cell(2, 12).value)
        address2 = self.driver.find_element_by_name(PracticeLocators.addressLine2)
        address2.clear()
        address2.send_keys(testData.cell(2, 13).value)
        addressCity = self.driver.find_element_by_name(PracticeLocators.addressCity)
        addressCity.clear()
        addressCity.send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        addressZipcode = self.driver.find_element_by_name(PracticeLocators.addressZipCode)
        addressZipcode.clear()
        addressZipcode.send_keys(testData.cell(2, 16).value)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_xpath(DispenserLocators.selectServiceType).click()
        time.sleep(1)
        self.driver.find_element_by_css_selector(DispenserLocators.courierService).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def OnetimeOTCProvidePayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.CreateOnetimeRXButton).click()
        self.driver.find_element_by_name(PracticeLocators.OnetimeSearchMedicine).send_keys(scriptData.cell(2, 2).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 2).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.providePayment).click()
        EditPayment = self.driver.find_elements(By.XPATH, DispenserLocators.editPaymentButton)
        if EditPayment:
            EditPayment[0].click()
        cardName = self.driver.find_element_by_name(PracticeLocators.cardName)
        cardName.clear()
        cardName.send_keys(testData.cell(2, 7).value)
        cardNumber = self.driver.find_element_by_name(PracticeLocators.maskCardNumberField)
        cardNumber.clear()
        cardNumber.send_keys(testData.cell(2, 8).value)
        cardCVV = self.driver.find_element_by_name(PracticeLocators.cardCVV)
        cardCVV.clear()
        cardCVV.send_keys(testData.cell(2, 9).value)
        Address1 = self.driver.find_element_by_name(PracticeLocators.addressLine1)
        Address1.clear()
        Address1.send_keys(testData.cell(2, 12).value)
        address2 = self.driver.find_element_by_name(PracticeLocators.addressLine2)
        address2.clear()
        address2.send_keys(testData.cell(2, 13).value)
        addressCity = self.driver.find_element_by_name(PracticeLocators.addressCity)
        addressCity.clear()
        addressCity.send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        addressZipcode = self.driver.find_element_by_name(PracticeLocators.addressZipCode)
        addressZipcode.clear()
        addressZipcode.send_keys(testData.cell(2, 16).value)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def OnetimeCompoundProvidePayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.CreateOnetimeRXButton).click()
        self.driver.find_element_by_name(PracticeLocators.OnetimeSearchMedicine).send_keys(scriptData.cell(2, 3).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 3).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.providePayment).click()
        EditPayment = self.driver.find_elements(By.XPATH, DispenserLocators.editPaymentButton)
        if EditPayment:
            EditPayment[0].click()
        cardName = self.driver.find_element_by_name(PracticeLocators.cardName)
        cardName.clear()
        cardName.send_keys(testData.cell(2, 7).value)
        cardNumber = self.driver.find_element_by_name(PracticeLocators.maskCardNumberField)
        cardNumber.clear()
        cardNumber.send_keys(testData.cell(2, 8).value)
        cardCVV = self.driver.find_element_by_name(PracticeLocators.cardCVV)
        cardCVV.clear()
        cardCVV.send_keys(testData.cell(2, 9).value)
        Address1 = self.driver.find_element_by_name(PracticeLocators.addressLine1)
        Address1.clear()
        Address1.send_keys(testData.cell(2, 12).value)
        address2 = self.driver.find_element_by_name(PracticeLocators.addressLine2)
        address2.clear()
        address2.send_keys(testData.cell(2, 13).value)
        addressCity = self.driver.find_element_by_name(PracticeLocators.addressCity)
        addressCity.clear()
        addressCity.send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        addressZipcode = self.driver.find_element_by_name(PracticeLocators.addressZipCode)
        addressZipcode.clear()
        addressZipcode.send_keys(testData.cell(2, 16).value)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def SubscriptionRXProvidePayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        time.sleep(2)
        el = self.driver.find_element_by_xpath(PracticeLocators.createSubscriptionRxButton)
        self.driver.execute_script('arguments[0].click()', el)
        self.driver.find_element_by_xpath(PracticeLocators.subscriptionProductSearch).send_keys(
            scriptData.cell(2, 1).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 1).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.providePayment).click()
        EditPayment = self.driver.find_elements(By.XPATH, DispenserLocators.editPaymentButton)
        if EditPayment:
            EditPayment[0].click()
        cardName = self.driver.find_element_by_name(PracticeLocators.cardName)
        cardName.clear()
        cardName.send_keys(testData.cell(2, 7).value)
        cardNumber = self.driver.find_element_by_name(PracticeLocators.maskCardNumberField)
        cardNumber.clear()
        cardNumber.send_keys(testData.cell(2, 8).value)
        cardCVV = self.driver.find_element_by_name(PracticeLocators.cardCVV)
        cardCVV.clear()
        cardCVV.send_keys(testData.cell(2, 9).value)
        Address1 = self.driver.find_element_by_name(PracticeLocators.addressLine1)
        Address1.clear()
        Address1.send_keys(testData.cell(2, 12).value)
        address2 = self.driver.find_element_by_name(PracticeLocators.addressLine2)
        address2.clear()
        address2.send_keys(testData.cell(2, 13).value)
        addressCity = self.driver.find_element_by_name(PracticeLocators.addressCity)
        addressCity.clear()
        addressCity.send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        addressZipcode = self.driver.find_element_by_name(PracticeLocators.addressZipCode)
        addressZipcode.clear()
        addressZipcode.send_keys(testData.cell(2, 16).value)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def SubscriptionOTCProvidePayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        time.sleep(2)
        el = self.driver.find_element_by_xpath(PracticeLocators.createSubscriptionRxButton)
        self.driver.execute_script('arguments[0].click()', el)
        self.driver.find_element_by_xpath(PracticeLocators.subscriptionProductSearch).send_keys(
            scriptData.cell(2, 2).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 2).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.providePayment).click()
        EditPayment = self.driver.find_elements(By.XPATH, DispenserLocators.editPaymentButton)
        if EditPayment:
            EditPayment[0].click()
        cardName = self.driver.find_element_by_name(PracticeLocators.cardName)
        cardName.clear()
        cardName.send_keys(testData.cell(2, 7).value)
        cardNumber = self.driver.find_element_by_name(PracticeLocators.maskCardNumberField)
        cardNumber.clear()
        cardNumber.send_keys(testData.cell(2, 8).value)
        cardCVV = self.driver.find_element_by_name(PracticeLocators.cardCVV)
        cardCVV.clear()
        cardCVV.send_keys(testData.cell(2, 9).value)
        Address1 = self.driver.find_element_by_name(PracticeLocators.addressLine1)
        Address1.clear()
        Address1.send_keys(testData.cell(2, 12).value)
        address2 = self.driver.find_element_by_name(PracticeLocators.addressLine2)
        address2.clear()
        address2.send_keys(testData.cell(2, 13).value)
        addressCity = self.driver.find_element_by_name(PracticeLocators.addressCity)
        addressCity.clear()
        addressCity.send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        addressZipcode = self.driver.find_element_by_name(PracticeLocators.addressZipCode)
        addressZipcode.clear()
        addressZipcode.send_keys(testData.cell(2, 16).value)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def SubscriptionCompoundProvidePayment(self):
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.CLASS_NAME, PracticeLocators.Add_RX))).click()
        PatientSearch = self.driver.find_element_by_xpath(PracticeLocators.Patient_search_textbox)
        PatientSearch.click()
        time.sleep(1)
        PatientSearch.send_keys(testData.cell(2, 1).value)
        self.driver.find_element_by_xpath("(//*[text()='" + testData.cell(2, 2).value + "'])[1]").click()
        DoctorSearch = self.driver.find_element_by_xpath(PracticeLocators.doctor_search_textbox)
        DoctorSearch.click()
        time.sleep(1)
        DoctorSearch.send_keys(scriptData.cell(2, 8).value)
        selectPractice = self.driver.find_elements(By.XPATH, "(//*[text()='" + scriptData.cell(2,
                                                                                               8).value + " " + "Practice'])[2]")
        if selectPractice:
            selectPractice[0].click()
        else:
            self.driver.find_element_by_xpath(
                "//*[text()='" + scriptData.cell(2, 8).value + " " + "Practice']").click()
        self.driver.find_element_by_xpath(PracticeLocators.NextButton).click()
        time.sleep(2)
        el = self.driver.find_element_by_xpath(PracticeLocators.createSubscriptionRxButton)
        self.driver.execute_script('arguments[0].click()', el)
        self.driver.find_element_by_xpath(PracticeLocators.subscriptionProductSearch).send_keys(
            scriptData.cell(2, 3).value)
        time.sleep(2)
        self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 3).value + "']").click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
        time.sleep(2)
        instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
        instruction.click()
        instruction.send_keys(testData.cell(2, 3).value)
        ExpiryDate = self.driver.find_elements(By.XPATH, PracticeLocators.surgeryDate)
        if ExpiryDate:
            ExpiryDate[0].send_keys(currentDate)
        self.driver.find_element_by_xpath(PracticeLocators.allergiesButton).click()
        allergies = self.driver.find_element_by_xpath(PracticeLocators.allergiesTextbox)
        allergies.clear()
        allergies.send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.doneButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.addDropchartButton).click()
        time.sleep(2)
        element = self.driver.find_element_by_xpath(PracticeLocators.selectDocuments)
        self.driver.execute_script("arguments[0].click()", element)
        self.driver.find_element_by_xpath(PracticeLocators.selectButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.providePayment).click()
        EditPayment = self.driver.find_elements(By.XPATH, DispenserLocators.editPaymentButton)
        if EditPayment:
            EditPayment[0].click()
        cardName = self.driver.find_element_by_name(PracticeLocators.cardName)
        cardName.clear()
        cardName.send_keys(testData.cell(2, 7).value)
        cardNumber = self.driver.find_element_by_name(PracticeLocators.maskCardNumberField)
        cardNumber.clear()
        cardNumber.send_keys(testData.cell(2, 8).value)
        cardCVV = self.driver.find_element_by_name(PracticeLocators.cardCVV)
        cardCVV.clear()
        cardCVV.send_keys(testData.cell(2, 9).value)
        Address1 = self.driver.find_element_by_name(PracticeLocators.addressLine1)
        Address1.clear()
        Address1.send_keys(testData.cell(2, 12).value)
        address2 = self.driver.find_element_by_name(PracticeLocators.addressLine2)
        address2.clear()
        address2.send_keys(testData.cell(2, 13).value)
        addressCity = self.driver.find_element_by_name(PracticeLocators.addressCity)
        addressCity.clear()
        addressCity.send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        addressZipcode = self.driver.find_element_by_name(PracticeLocators.addressZipCode)
        addressZipcode.clear()
        addressZipcode.send_keys(testData.cell(2, 16).value)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_xpath(PracticeLocators.surgeryDate).send_keys(currentDate)
        Notes = self.driver.find_elements(By.XPATH, PracticeLocators.pharmacyNotes)
        if Notes:
            Notes[0].send_keys(testData.cell(2, 5).value)
        else:
            self.driver.find_element_by_xpath(DispenserLocators.addNoteButton).click()
            self.driver.find_element_by_xpath(DispenserLocators.orderNoteTextbox).send_keys(
                testData.cell(2, 5).value)
            pharmaCheckbox = self.driver.find_elements(By.XPATH, DispenserLocators.pharmacyCheckbox)
            if pharmaCheckbox:
                pharmaCheckbox[0].click()
            self.driver.find_element_by_xpath(PatientLocators.iframeSaveButton).click()
        createOrderButton = self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton)
        self.driver.execute_script("arguments[0].click()", createOrderButton)
        time.sleep(5)
        order_id = None
        for request in self.driver.requests:
            if request.response:
                if request.method == 'POST' and (
                        request.url.__contains__('/orders/') or request.url.__contains__('/subscriptions/')):
                    # print(request.method + ' ' + request.url)
                    try:
                        data = json.loads(request.response.body)
                        # print('parsed as json')
                        if '_id' in data:
                            order_id = data['_id']
                    except UnicodeDecodeError:
                        try:
                            data = json.loads(decompress(request.response.body))
                            # print('decompressed and parsed as json')
                            if '_id' in data:
                                order_id = data['_id']
                        except json.decoder.JSONDecodeError:
                            data = request.response.body
                            # print('decompressed and parsed as string')
                    # print(data)
        print(order_id)
        scriptData.cell(2, 13).value = order_id
        datafile.save(FilePath)

    def VerifyOrderDetailsScreen(self):
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Orders'))).click()
        time.sleep(4)
        self.driver.find_element_by_xpath('//table/tbody/tr[1]/td[2]/a').click()
        time.sleep(3)
        assert 'Prescription & Order Details'
        time.sleep(1)
        assert self.driver.current_url.__contains__(scriptData.cell(2, 13).value)

    def EditOrder(self):
        self.driver.get(loginData.cell(3, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Edit Contents'))).click()
        self.driver.find_element_by_xpath(PracticeLocators.searchForMedication).send_keys('Compound')
        time.sleep(3)
        self.driver.find_element_by_xpath('//div[@data-test="_HINTS"]/div[2]').click()
        self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
        time.sleep(2)
        self.driver.find_element_by_xpath('//*[@class="sc-epnACN hXOwfY"]').click()
        self.driver.find_element_by_xpath(PracticeLocators.skipPayment).click()
        self.driver.find_element_by_xpath(PracticeLocators.updateOrderButton).click()
        time.sleep(7)
        if "You're all set!" in self.driver.page_source:
            self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
        else:
            print("\nOrder not created")
        self.driver.find_element_by_class_name('account-link-label').click()
        self.driver.find_element_by_link_text('Sign Out').click()

    def EditPatientDetails(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.editButton))).click()
        cardName = self.driver.find_element_by_name(PracticeLocators.cardName)
        cardName.clear()
        cardName.send_keys(testData.cell(2, 7).value)
        cardNumber = self.driver.find_element_by_name(PracticeLocators.maskCardNumberField)
        cardNumber.clear()
        cardNumber.send_keys(testData.cell(2, 8).value)
        cardCVV = self.driver.find_element_by_name(PracticeLocators.cardCVV)
        cardCVV.clear()
        cardCVV.send_keys(testData.cell(2, 9).value)
        self.driver.find_element_by_name(PracticeLocators.addressLine1).send_keys(testData.cell(2, 12).value)
        self.driver.find_element_by_name(PracticeLocators.addressLine2).send_keys(testData.cell(2, 13).value)
        self.driver.find_element_by_name(PracticeLocators.addressCity).send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        self.driver.find_element_by_name(PracticeLocators.addressZipCode).send_keys(testData.cell(2, 16).value)
        self.driver.find_element_by_xpath(DispenserLocators.saveAndExitButton).click()
        self.driver.implicitly_wait(10)
        assert "Order has been successfully updated."
        button = WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, PracticeLocators.dismissButton)))
        self.driver.execute_script("arguments[0].click()", button)

    def CompleteOrder(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.selectActionDropdown))).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(DispenserLocators.completeValue).click()
        self.driver.find_element_by_class_name('form-control').send_keys('Patient completed the treatment.')
        self.driver.find_element_by_xpath(PracticeLocators.submitButton).click()
        time.sleep(2)
        assert "You have successfully completed this order"
        button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", button)

    def CancelOrder(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.selectActionDropdown))).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(DispenserLocators.cancelValue).click()
        self.driver.find_element_by_class_name('form-control').send_keys('I do not need this order.')
        self.driver.find_element_by_xpath(PracticeLocators.submitButton).click()
        time.sleep(2)
        assert "You have successfully canceled this order"
        button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", button)

    def SendOutOfNetwork(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.selectActionDropdown))).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(DispenserLocators.sendOutOfNetwork).click()
        self.driver.find_element_by_class_name('form-control').send_keys('Order send out of network.')
        self.driver.find_element_by_xpath(PracticeLocators.submitButton).click()
        time.sleep(2)
        assert "You have successfully sent this order out of network"
        button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", button)

    def ApproveOrderFromPractice(self):
        assert self.driver.find_element_by_xpath(DispenserLocators.verifyOrderDate)
        assert self.driver.find_element_by_xpath(
            "(//tr[@class='table-row'])[1]//td[3][text()='" + testData.cell(2, 1).value + "']")
        self.driver.find_element_by_xpath(DispenserLocators.approveButton).click()
        time.sleep(2)
        assert "Please confirm that you approve the selected prescriptions."
        time.sleep(1)
        self.driver.find_element_by_xpath(DispenserLocators.yesButton).click()
        time.sleep(1)
        assert "You have successfully approved the selected prescriptions!"
        self.driver.find_element_by_xpath(PracticeLocators.closeButton).click()

    def PatientApprovalAndTransfer(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.confirmApprovalButton))).click()
        time.sleep(1)
        assert "Patient payment approval has been sucessfully captured."
        dismissButton = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", dismissButton)
        self.driver.find_element_by_xpath(DispenserLocators.transferOrderButton).click()
        time.sleep(1)
        assert "Have you attached the proper document(s) for this order?"
        self.driver.find_element_by_xpath(DispenserLocators.yesButton).click()
        time.sleep(1)
        assert "This order has been successfully transferred."
        time.sleep(2)
        Button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", Button)

    def ProcessPaymentAndCreateLable(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.processPaymentButton))).click()
        time.sleep(3)
        assert "Congratulations, payment has been successfully processed!"
        self.driver.find_element_by_xpath(DispenserLocators.createPostageLabelButton).click()
        self.driver.implicitly_wait(10)
        assert "(Payment processed once label is created)"
        self.driver.find_element_by_xpath(DispenserLocators.createLabelButton).click()
        time.sleep(4)
        assert "Congratulations, you have successfully created a postage label."
        self.driver.find_element_by_xpath(DispenserLocators.printLabelButton).click()
        time.sleep(1)

    def ProcessPaymentAndConfirmPickUpPerson(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.processPaymentButton))).click()
        time.sleep(3)
        assert "Congratulations, payment has been successfully processed!"
        self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
        self.driver.find_element_by_xpath(DispenserLocators.confirmPickUpButton).click()
        time.sleep(2)
        assert "You have successfully completed this order"
        dismissButton = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", dismissButton)

    def ProcessPaymentAndConfirmCourierPickUp(self):
        self.driver.get(loginData.cell(4, 2).value + "/orders/" + scriptData.cell(2, 13).value)
        WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, DispenserLocators.processPaymentButton))).click()
        time.sleep(3)
        assert "Congratulations, payment has been successfully processed!"
        self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
        self.driver.find_element_by_xpath(DispenserLocators.confirmCourierPickUp).click()
        time.sleep(1)
        assert "Order successfully updated."
        dismissButton = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", dismissButton)
        self.driver.find_element_by_xpath(DispenserLocators.completeButton).click()
        time.sleep(1)
        assert "You have successfully completed this order"
        btn = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", btn)
