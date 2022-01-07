from openpyxl import load_workbook

FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
loginData = datafile["Login Credentials"]


class ReadConfig:
    @staticmethod
    def getApplicationURL():
        url = loginData.cell(2, 2).value
        return url

    @staticmethod
    def getUsername():
        username = loginData.cell(2, 3).value
        return username

    @staticmethod
    def getPassword():
        password = loginData.cell(2, 4).value
        return password
