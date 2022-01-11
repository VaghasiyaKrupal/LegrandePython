import time
from datetime import date
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from Locators.DispenserLocators import DispenserLocators
from Locators.Locators import Locators
from Locators.PatientLocators import PatientLocators
from Locators.PracticeLocators import PracticeLocators
import random
from faker import Faker

FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']
loginData = datafile['Login Credentials']

number = random.randint(0, 99999)
faker = Faker()
today = date.today()
currentDate = today.strftime("%m/%d/%Y")
FirstName = faker.first_name()
LastName = faker.last_name()
PhoneNumber = faker.phone_number()
street_1 = faker.address()
street_2 = faker.address()
city = faker.city()
zipCode = faker.zipcode()
acNumber = random.randint(10, 50)


class TestPatient:
    def test_CreatePatient(self, setup, PracticeLogin, CreatePatient):
        self.driver = setup

    def test_CreatingCompoundSubscriptionSkipPayment(self, setup, PracticeLogin, SubscriptionCompoundSkipPayment):
        self.driver = setup
        if "You're all set!" in self.driver.page_source:
            self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
            print('\nOrder: Subscription Compound with Skip payment, created successfully')
        else:
            print("\nOrder: Subscription Compound with Skip payment, is not created")
        self.driver.close()

    def test_CreateMailGetRegistrationLink(self, setup):
        self.driver = setup
        self.driver.get(Locators.malinatorLink)
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_id(Locators.searchBoxID).send_keys(loginData.cell(8, 3).value)
        self.driver.find_element_by_xpath(Locators.goButtonLink).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(Locators.emailInbox).click()
        time.sleep(3)
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, Locators.emailIframe)))
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, Locators.completeAccount))).click()
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_id(PatientLocators.newPasswordID).send_keys(loginData.cell(2, 4).value)
        self.driver.find_element_by_id(PatientLocators.confirmPasswordID).send_keys(loginData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.submit_CreateOrderButton).click()
        time.sleep(3)
        if "Thank you!" and "Your email address has been verified, and your registration is complete." in self.driver.page_source:
            print('\nYour patient account has been verified')
        else:
            print('\nPatient account has not verified!')
        self.driver.quit()

    def test_AccountSetup(self, setup, PatientLogin):
        self.driver = setup
        self.driver.find_element_by_id(PatientLocators.allergies).send_keys(testData.cell(2, 4).value)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        time.sleep(2)
        self.driver.find_element_by_id(PatientLocators.payorAccountNumber).send_keys(number)
        self.driver.find_element_by_id(PatientLocators.payorGroupNumber).send_keys(number)
        self.driver.find_element_by_id(PatientLocators.payorRxBinNumber).send_keys(number)
        self.driver.find_element_by_id(PatientLocators.payorPCnNumber).send_keys(number)
        self.driver.find_element_by_id(PatientLocators.payorPhoneNumber).send_keys(PhoneNumber)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        time.sleep(2)
        if testData.cell(2, 1).value in self.driver.page_source:
            print('\nAccount setup is completed')
        else:
            print('\nYour were logged in as unknown user')
        self.driver.quit()

    def test_ForgotPassword(self, setup, PatientForgotPassword):
        self.driver = setup
        time.sleep(3)
        if "Password Email Sent." and "Please check your email for a link to reset your password." in self.driver.page_source:
            print('\nPassword Reset Email Sent to this email address: '+loginData.cell(8, 3).value)
        else:
            print("\nOpps, Something went wrong")
        self.driver.quit()

    def test_CheckOrderDetails(self, setup, PatientLogin):
        self.driver = setup
        self.driver.find_element_by_link_text('orders').click()
        self.driver.implicitly_wait(5)
        self.driver.find_element_by_class_name(PatientLocators.firstOrder).click()
        time.sleep(2)
        assert "Tracking Details" in self.driver.page_source
        assert "Shipment" in self.driver.page_source
        assert "Payment Details" in self.driver.page_source
        assert "Billing Address" in self.driver.page_source
        assert "Shipping Address" in self.driver.page_source
        if self.driver.find_element_by_id('contact-help'):
            print('Footer is present')
        else:
            print('Footer is not available')
        self.driver.find_element_by_xpath('//*[text()="Order Details"]').click()
        self.driver.quit()

    def test_EditPlan(self, setup, PatientLogin):
        self.driver = setup
        self.driver.find_element_by_link_text('plans').click()
        self.driver.find_element_by_class_name('css-q8kuvx').click()
        self.driver.find_element_by_xpath("//h2[contains(text(),'Manage Your Plan')]").click()
        self.driver.find_element_by_xpath('//input[@placeholder="MM/DD/YYYY"]').click()
        self.driver.find_element_by_xpath(PatientLocators.newDate).click()
        self.driver.find_element_by_xpath(DispenserLocators.updateButton).click()
        time.sleep(2)
        assert "You have successfully saved your updates." in self.driver.page_source
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.quit()

    def test_CancelPlan(self, setup, PatientLogin):
        self.driver = setup
        self.driver.find_element_by_link_text('plans').click()
        self.driver.find_element_by_class_name('css-q8kuvx').click()
        self.driver.find_element_by_xpath("//h2[contains(text(),'Manage Your Plan')]").click()
        self.driver.find_element_by_xpath(PatientLocators.cancelPlanCheckbox).click()
        # time.sleep(1)
        self.driver.find_element_by_xpath(DispenserLocators.updateButton).click()
        # time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        # time.sleep(1)
        self.driver.find_element_by_id(PatientLocators.reasoneDropdownID).click()
        self.driver.find_element_by_xpath(PatientLocators.selectOtherReason).click()
        self.driver.find_element_by_id(PatientLocators.userReasonID).send_keys('I have completed my treatment.')
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.submitButton)
        time.sleep(2)
        assert "Your plan has been canceled." in self.driver.page_source
        assert "Thank you for your feedback!" in self.driver.page_source


class TestSetting:
    def test_AccountInfo(self, setup, PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url+"/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientAccount).click()
        time.sleep(2)
        assert 'Contact Details' and 'Payment Method' and 'Shipping Address' in self.driver.page_source
        self.driver.find_element_by_xpath('(//*[text()="Update"])[1]').click()
        fName = self.driver.find_element_by_name(PracticeLocators.firstName)
        fName.send_keys(Keys.CONTROL+"a")
        fName.send_keys(Keys.DELETE)
        fName.send_keys(FirstName)
        lName = self.driver.find_element_by_name(PracticeLocators.lastName)
        lName.send_keys(Keys.CONTROL+"a")
        lName.send_keys(Keys.DELETE)
        lName.send_keys(LastName)
        phone = self.driver.find_element_by_name(PracticeLocators.phoneNumber)
        phone.send_keys(Keys.CONTROL+"a")
        phone.send_keys(Keys.DELETE)
        phone.send_keys(PhoneNumber)
        self.driver.find_element_by_xpath(PatientLocators.saveUpdateButton).click()
        time.sleep(2)
        if "You have successfully saved your updates." in self.driver.page_source:
            self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
            print("Patient details updated successfully")
        else:
            print("\nPatient details not updated")
        self.driver.quit()

    def test_PaymentDetails(self, setup, PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientAccount).click()
        time.sleep(2)
        assert 'Contact Details' and 'Payment Method' and 'Shipping Address' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.updatePaymentButton).click()
        card = self.driver.find_element_by_id(PracticeLocators.cardName)
        card.send_keys(Keys.CONTROL+'a')
        card.send_keys(Keys.DELETE)
        card.send_keys("Master Card")
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[contains(@src,'hpc.uat.freedompay.com')]")))
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLogin.CardNumber))).send_keys(PatientLocators.cardDetails)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLocators.ExperientionDate))).send_keys(PatientLocators.cardExpiryDate)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLocators.cvvCode))).send_keys(PatientLocators.cardCvvCode)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLocators.PostalCode))).send_keys(PatientLocators.cardPostalCode)
        time.sleep(1)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, PatientLocators.iframeSaveButton))).click()
        # time.sleep(5)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath('//section[@role="dialog"]//*[text()="Continue"]').click()
        assert self.driver.find_element_by_xpath('//*[@src="/patient/mastercard.png"]')
        self.driver.quit()

    def test_ShippingAddress(self, setup, PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientAccount).click()
        time.sleep(2)
        assert 'Contact Details' and 'Payment Method' and 'Shipping Address' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.updateShippingButton).click()

        address1 = self.driver.find_element_by_name(PracticeLocators.addressLine1)
        address1.send_keys(Keys.CONTROL+'a')
        address1.send_keys(Keys.DELETE)
        address1.send_keys(street_1)

        City = self.driver.find_element_by_name(PracticeLocators.addressCity)
        City.send_keys(Keys.CONTROL + 'a')
        City.send_keys(Keys.DELETE)
        City.send_keys(city)

        self.driver.find_element_by_xpath(PatientLocators.shippingState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PatientLocators.stateValue).click()

        zip = self.driver.find_element_by_name(PracticeLocators.addressZipCode)
        zip.send_keys(Keys.CONTROL+'a')
        zip.send_keys(Keys.DELETE)
        zip.send_keys(zipCode)
        time.sleep(1)

        self.driver.find_element_by_xpath(PatientLocators.saveUpdateButton).click()
        time.sleep(2)
        assert "You have successfully saved your updates." in self.driver.page_source
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.quit()

    def test_ChangePassword(self, setup, PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientPassword).click()
        self.driver.find_element_by_id(PatientLocators.currentPasswordID).send_keys(loginData.cell(2, 4).value)
        self.driver.find_element_by_id(PatientLocators.newPasswordID).send_keys(loginData.cell(2, 4).value)
        self.driver.find_element_by_id(PatientLocators.confirmPasswordID).send_keys(loginData.cell(2, 4).value)
        self.driver.find_element_by_xpath(DispenserLocators.updateButton).click()
        self.driver.quit()

    def test_AddInsurance(self, setup, PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientHealthInsurance).click()
        time.sleep(2)
        if 'You do not currently have any Health Insurance saved. Select "Add New Insurance" to save your Health Insurance' in self.driver.page_source:
            self.driver.find_element_by_xpath(PatientLocators.addNewInsurance).click()
            time.sleep(3)
            self.driver.find_element_by_id(PatientLocators.payorID).click()
            self.driver.find_element_by_xpath(PatientLocators.firstPayor).click()
            self.driver.find_element_by_id(PatientLocators.payorAccountNumber).send_keys(acNumber)
            self.driver.find_element_by_id(PatientLocators.payorGroupNumber).send_keys(acNumber)
            self.driver.find_element_by_id(PatientLocators.payorRxBinNumber).send_keys(acNumber)
            self.driver.find_element_by_id(PatientLocators.payorPCnNumber).send_keys(acNumber)
            self.driver.find_element_by_id(PatientLocators.payorPhoneNumber).send_keys(PhoneNumber)
            time.sleep(1)
            self.driver.find_element_by_xpath(PatientLocators.addInsuranceButton).click()
            time.sleep(2)
            assert "You have sucessfully added a new health insurance!" in self.driver.page_source
            self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        else:
            self.driver.find_element_by_link_text('Edit').click()
            time.sleep(2)
            self.driver.find_element_by_id(PatientLocators.payorID).click()
            self.driver.find_element_by_xpath(PatientLocators.editPayorName).click()
            self.driver.find_element_by_xpath(PatientLocators.saveUpdateButton).click()
            time.sleep(2)
            assert "You have successfully saved your updates." in self.driver.page_source
            self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.quit()

    def test_RemoveInsurance(self, setup,PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientHealthInsurance).click()
        time.sleep(2)
        if "Remove" in self.driver.page_source:
            self.driver.find_element_by_xpath('//*[text()="Remove"]').click()
            assert "Are you sure you want to delete this insurance?" in self.driver.page_source
            self.driver.find_element_by_xpath(PatientLocators.confirmButton).click()
        else:
            print("\nYou do not have any insurance")
        self.driver.quit()

    def test_AddAllergies(self, setup, PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientAllergies).click()
        time.sleep(1)
        if 'You do not currently have any allergies saved. Enter an allergy in the field and select "Add Allergy" to add it to your profile.' in self.driver.page_source:
            self.driver.find_element_by_id(PatientLocators.allergiesID).send_keys(FirstName)
            self.driver.find_element_by_xpath(PatientLocators.addAllergyButton).click()
        else:
            self.driver.find_element_by_id(PatientLocators.allergiesID).send_keys(FirstName)
            self.driver.find_element_by_xpath(PatientLocators.addAllergyButton).click()
        self.driver.quit()

    def test_RemoveAllergies(self, setup, PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientAllergies).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PatientLocators.removeAllergyButton).click()
        self.driver.find_element_by_xpath(PatientLocators.confirmButton).click()
        self.driver.quit()

    def testCheckPreference(self, setup,PatientLogin):
        self.driver = setup
        url = self.driver.current_url
        newURl = url + "/settings"
        self.driver.get(newURl)
        time.sleep(2)
        assert 'Account' and 'Password' and 'Health Insurance' and 'Allergies' and 'Preference' in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.patientPreference).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PatientLocators.messageCheckbox).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PatientLocators.additionalCheckbox).click()
        self.driver.quit()


class TestEnterDetailsFromNotification:
    def test_CompleteDetailsFromNotification(self, setup, PatientLogin):
        self.driver = setup
        time.sleep(3)
        assert 'Your order is ready!' and "We just need a few more details from you, then we'll be ready to ship!" in self.driver.page_source
        self.driver.find_element_by_xpath(PatientLocators.completePayment).click()
        self.driver.find_element_by_name(PracticeLocators.addressLine1).send_keys(street_1)
        self.driver.find_element_by_name(PracticeLocators.addressCity).send_keys(city)
        self.driver.find_element_by_xpath(PatientLocators.shippingState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PatientLocators.stateValue).click()
        self.driver.find_element_by_name(PracticeLocators.addressZipCode).send_keys(zipCode)
        time.sleep(1)

        WebDriverWait(self.driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[contains(@src,'hpc.uat.freedompay.com')]")))
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLocators.CardNumber))).send_keys(
            PatientLocators.cardDetails)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLocators.ExperientionDate))).send_keys(
            PatientLocators.cardExpiryDate)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLocators.cvvCode))).send_keys(
            PatientLocators.cardCvvCode)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, PatientLocators.PostalCode))).send_keys(
            PatientLocators.cardPostalCode)
        time.sleep(1)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, PatientLocators.iframeSaveButton))).click()
        # time.sleep(5)
        self.driver.switch_to.default_content()
        time.sleep(4)
        assert "Thank You!" in self.driver.page_source
        self.driver.quit()

