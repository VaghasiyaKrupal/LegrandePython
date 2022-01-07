import time
from openpyxl import load_workbook
from Locators.PatientLocators import PatientLocators
from Locators.PracticeLocators import PracticeLocators
from Locators.Locators import Locators

FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']


class TestStandardPharma:
    def test_ForgotPassword(self, setup, RXPharmaForgotPassword):
        self.driver = setup
        self.driver.close()

    def test_CreatePatient(self, setup, RXPharmaLogin, CreatePatient):
        self.driver = setup

    def test_CreateUser(self, setup, RXPharmaLogin, CreateUser):
        self.driver = setup

    def test_PatientMainSearch(self, setup, RXPharmaLogin, PatientMainSearch):
        self.driver = setup

    def test_ChangeProductPrice(self, setup, RXPharmaLogin, ChangeProductPrice):
        self.driver = setup

    def test_CreatingRXOnetimeSkipPayment(self, setup, RXPharmaLogin, OnetimeRXSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_ApproveOnetimeFromPracticeAccount(self, setup, PracticeLogin):
        self.driver = setup
        assert self.driver.find_element_by_xpath(Locators.verifyOrderDate)
        assert self.driver.find_element_by_xpath(Locators.verifyPatientName)
        self.driver.find_element_by_xpath(Locators.approveButton).click()
        time.sleep(2)
        assert "Please confirm that you approve the selected prescriptions."
        time.sleep(1)
        self.driver.find_element_by_xpath(Locators.yesButton).click()
        time.sleep(1)
        assert "You have successfully approved the selected prescriptions!"
        self.driver.find_element_by_xpath(PracticeLocators.closeButton).click()
        self.driver.close()

    def test_ProcessPaymentForOnetimeAndCreateLabel(self, setup, RXPharmaLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(Locators.processPaymentButton).click()
        time.sleep(3)
        assert "Congratulations, payment has been successfully processed! Click below to create a postage label."
        self.driver.find_element_by_xpath(Locators.createPostageLabelButton).click()
        self.driver.implicitly_wait(10)
        assert "(Payment processed once label is created)"
        self.driver.find_element_by_xpath(Locators.createLabelButton).click()
        time.sleep(3)
        assert "Congratulations, you have successfully created a postage label."
        self.driver.find_element_by_xpath(Locators.printLabelButton).click()
        time.sleep(1)
        self.driver.quit()

    def test_CreatingOTCOnetimeSkipPayment(self, setup, RXPharmaLogin, OnetimeOTCSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_CreatingCompoundOnetimeSkipPayment(self, setup, RXPharmaLogin, OnetimeCompoundSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_CreatingRXSubscriptionSkipPayment(self, setup, RXPharmaLogin, SubscriptionRXSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_ApproveSubscriptionFromPracticeAccount(self, setup, PracticeLogin):
        self.driver = setup
        assert self.driver.find_element_by_xpath(Locators.verifyOrderDate)
        assert self.driver.find_element_by_xpath(Locators.verifyPatientName)
        self.driver.find_element_by_xpath(Locators.approveButton).click()
        time.sleep(2)
        assert "Please confirm that you approve the selected prescriptions."
        time.sleep(1)
        self.driver.find_element_by_xpath(Locators.yesButton).click()
        time.sleep(1)
        assert "You have successfully approved the selected prescriptions!"
        self.driver.find_element_by_xpath(PracticeLocators.closeButton).click()
        self.driver.close()

    def test_ProcessPaymentForSubscriptionAndCreateLabel(self, setup, RXPharmaLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(Locators.processPaymentButton).click()
        time.sleep(3)
        assert "Congratulations, payment has been successfully processed! Click below to create a postage label."
        self.driver.find_element_by_xpath(Locators.createPostageLabelButton).click()
        self.driver.implicitly_wait(10)
        assert "(Payment processed once label is created)"
        self.driver.find_element_by_xpath(Locators.createLabelButton).click()
        time.sleep(3)
        assert "Congratulations, you have successfully created a postage label."
        self.driver.find_element_by_xpath(Locators.printLabelButton).click()
        time.sleep(1)
        self.driver.quit()

    def test_CreatingOTCSubscriptionSkipPayment(self, setup, RXPharmaLogin, SubscriptionOTCSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_CreatingCompoundSubscriptionSkipPayment(self, setup, RXPharmaLogin, SubscriptionCompoundSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_RXOnetimeProvidePayment(self, setup, RXPharmaLogin, OnetimeRXProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_ApprovePaymentOptionOnetimeOrderFromPracticeAccount(self, setup, PracticeLogin):
        self.driver = setup
        assert self.driver.find_element_by_xpath(Locators.verifyOrderDate)
        assert self.driver.find_element_by_xpath(Locators.verifyPatientName)
        self.driver.find_element_by_xpath(Locators.approveButton).click()
        time.sleep(2)
        assert "Please confirm that you approve the selected prescriptions."
        time.sleep(1)
        self.driver.find_element_by_xpath(Locators.yesButton).click()
        time.sleep(1)
        assert "You have successfully approved the selected prescriptions!"
        self.driver.find_element_by_xpath(PracticeLocators.closeButton).click()
        self.driver.close()

    def test_ProcessPaymentForPaymentOptionOrderOnetimeAndCreateLabel(self, setup, RXPharmaLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(Locators.processPaymentButton).click()
        time.sleep(3)
        assert "Congratulations, payment has been successfully processed! Click below to create a postage label."
        self.driver.find_element_by_xpath(Locators.createPostageLabelButton).click()
        self.driver.implicitly_wait(10)
        assert "(Payment processed once label is created)"
        self.driver.find_element_by_xpath(Locators.createLabelButton).click()
        time.sleep(3)
        assert "Congratulations, you have successfully created a postage label."
        self.driver.find_element_by_xpath(Locators.printLabelButton).click()
        time.sleep(1)
        self.driver.quit()

    def test_OTCOnetimeProvidePayment(self, setup, RXPharmaLogin, OnetimeOTCProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_CompoundOnetimeProvidePayment(self, setup, RXPharmaLogin, OnetimeCompoundProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_RXSubscriptionProvidePayment(self, setup, RXPharmaLogin, SubscriptionRXProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_ApprovePaymentOptionSubscriptionOrderFromPracticeAccount(self, setup, PracticeLogin):
        self.driver = setup
        assert self.driver.find_element_by_xpath(Locators.verifyOrderDate)
        assert self.driver.find_element_by_xpath(Locators.verifyPatientName)
        self.driver.find_element_by_xpath(Locators.approveButton).click()
        time.sleep(2)
        assert "Please confirm that you approve the selected prescriptions."
        time.sleep(1)
        self.driver.find_element_by_xpath(Locators.yesButton).click()
        time.sleep(1)
        assert "You have successfully approved the selected prescriptions!"
        self.driver.find_element_by_xpath(PracticeLocators.closeButton).click()
        self.driver.close()

    def test_ProcessPaymentForPaymentOptionOrderSubscriptionAndCreateLabel(self, setup, RXPharmaLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(Locators.processPaymentButton).click()
        time.sleep(3)
        assert "Congratulations, payment has been successfully processed! Click below to create a postage label."
        self.driver.find_element_by_xpath(Locators.createPostageLabelButton).click()
        self.driver.implicitly_wait(10)
        assert "(Payment processed once label is created)"
        self.driver.find_element_by_xpath(Locators.createLabelButton).click()
        time.sleep(3)
        assert "Congratulations, you have successfully created a postage label."
        self.driver.find_element_by_xpath(Locators.printLabelButton).click()
        time.sleep(1)
        self.driver.quit()

    def test_OTCSubscriptionProvidePayment(self, setup, RXPharmaLogin, SubscriptionOTCProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_CompoundSubscriptionProvidePayment(self, setup, RXPharmaLogin, SubscriptionCompoundProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_EditOrder(self, setup, RXPharmaLogin, EditOrder):
        self.driver = setup
