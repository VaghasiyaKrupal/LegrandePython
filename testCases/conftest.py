import time
import pytest
from faker import Faker
from datetime import date
from seleniumwire import webdriver
from openpyxl import load_workbook
from Locators.Locators import Locators
from selenium.webdriver.common.by import By
from pageObjects.BaseFile import CommanFlow
from pageObjects.LoginPage import LoginScreen
from selenium.webdriver.chrome.options import Options
from Locators.PracticeLocators import PracticeLocators
from webdriver_manager.chrome import ChromeDriverManager

faker = Faker()
today = date.today()
currentDate = today.strftime("%m/%d/%Y")
FirstName = faker.first_name()
LastName = faker.last_name()
PhoneNumber = faker.phone_number()
EmailAddress = FirstName + "." + LastName + "@mailinator.com"

FilePath = "TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']
loginData = datafile["Login Credentials"]
scriptData = datafile['Script Data']


@pytest.fixture()
def setup():
    global driver
    mode = Options()
    mode.add_argument("--headless")
    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=mode)
    driver.maximize_window()
    return driver


@pytest.fixture()
def MasterLogin(setup):
    driver.get(loginData.cell(2, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(2, 3).value)
    login.SetPassword(loginData.cell(2, 4).value)
    login.SignIn()
    driver.implicitly_wait(10)


@pytest.fixture()
def MasterForgotPassword(setup):
    driver.get(loginData.cell(2, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.ForgotPassword(loginData.cell(2, 3).value)
    time.sleep(1)
    login.SubmitButton()


@pytest.fixture()
def PracticeLogin(setup):
    driver.get(loginData.cell(3, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(3, 3).value)
    login.SetPassword(loginData.cell(3, 4).value)
    login.SignIn()
    checkboxes = driver.find_elements(By.XPATH, Locators.termsCheckbox)
    if checkboxes:
        checkboxes[0].click()
        driver.find_element(By.XPATH, PracticeLocators.continueButton).click()


@pytest.fixture()
def PracticeForgotPassword(setup):
    driver.get(loginData.cell(3, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.ForgotPassword(loginData.cell(3, 3).value)
    time.sleep(1)
    login.SubmitButton()


@pytest.fixture()
def HubLogin(setup):
    driver.get(loginData.cell(4, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(4, 3).value)
    login.SetPassword(loginData.cell(4, 4).value)
    login.SignIn()
    checkboxes = driver.find_elements(By.XPATH, Locators.termsCheckbox)
    if checkboxes:
        checkboxes[0].click()
        driver.find_element(By.XPATH, PracticeLocators.continueButton).click()


@pytest.fixture()
def HubForgotPassword(setup):
    driver.get(loginData.cell(4, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.ForgotPassword(loginData.cell(4, 3).value)
    time.sleep(1)
    login.SubmitButton()


@pytest.fixture()
def RXPharmaLogin(setup):
    driver.get(loginData.cell(5, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(5, 3).value)
    login.SetPassword(loginData.cell(5, 4).value)
    login.SignIn()
    checkboxes = driver.find_elements(By.XPATH, Locators.termsCheckbox)
    if checkboxes:
        checkboxes[0].click()
        driver.find_element(By.XPATH, PracticeLocators.continueButton).click()


@pytest.fixture()
def RXPharmaForgotPassword(setup):
    driver.get(loginData.cell(5, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.ForgotPassword(loginData.cell(5, 3).value)
    time.sleep(1)
    login.SubmitButton()


@pytest.fixture()
def OTCPharmaLogin(setup):
    driver.get(loginData.cell(6, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(6, 3).value)
    login.SetPassword(loginData.cell(6, 4).value)
    login.SignIn()
    checkboxes = driver.find_elements(By.XPATH, Locators.termsCheckbox)
    if checkboxes:
        checkboxes[0].click()
        driver.find_element(By.XPATH, PracticeLocators.continueButton).click()


@pytest.fixture()
def OTCPharmaForgotPassword(setup):
    driver.get(loginData.cell(6, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.ForgotPassword(loginData.cell(6, 3).value)
    time.sleep(1)
    login.SubmitButton()


@pytest.fixture()
def CompoundPharmaLogin(setup):
    driver.get(loginData.cell(7, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(7, 3).value)
    login.SetPassword(loginData.cell(7, 4).value)
    login.SignIn()
    checkboxes = driver.find_elements(By.XPATH, Locators.termsCheckbox)
    if checkboxes:
        checkboxes[0].click()
        driver.find_element(By.XPATH, PracticeLocators.continueButton).click()


@pytest.fixture()
def CompoundPharmaForgotPassword(setup):
    driver.get(loginData.cell(7, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.ForgotPassword(loginData.cell(7, 3).value)
    time.sleep(1)
    login.SubmitButton()


@pytest.fixture()
def PatientLogin(setup):
    driver.get(loginData.cell(8, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(8, 3).value)
    login.SetPassword(loginData.cell(8, 4).value)
    login.SignIn()
    time.sleep(3)


@pytest.fixture()
def PatientForgotPassword(setup):
    driver.get(loginData.cell(8, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.PatientForgotPassword(loginData.cell(3, 3).value)
    time.sleep(1)
    login.SendResetEmail()


@pytest.fixture()
def UserLogin(setup):
    driver.get(loginData.cell(9, 2).value)
    driver.implicitly_wait(5)
    login = LoginScreen(driver)
    login.SetUsername(loginData.cell(9, 3).value)
    login.SetPassword(loginData.cell(9, 4).value)
    login.SignIn()
    checkboxes = driver.find_elements(By.XPATH, Locators.termsCheckbox)
    if checkboxes:
        checkboxes[0].click()
        driver.find_element(By.XPATH, PracticeLocators.continueButton).click()


@pytest.fixture()
def CreatePatient(setup):
    patient = CommanFlow(driver)
    patient.CreatePatient()


@pytest.fixture()
def CreateUser(setup):
    user = CommanFlow(driver)
    user.CreateUser()


@pytest.fixture()
def PatientMainSearch(setup):
    mainSearch = CommanFlow(driver)
    mainSearch.PatientMainSearch()


@pytest.fixture()
def ChangeProductPrice(setup):
    product = CommanFlow(driver)
    product.ChangeProductPrice()


@pytest.fixture()
def OnetimeRXSkipPayment(setup):
    onetimeRxSkipPayment = CommanFlow(driver)
    onetimeRxSkipPayment.OnetimeRXSkipPayment()


@pytest.fixture()
def OnetimeOTCSkipPayment(setup):
    onetimeOTCSkipPayment = CommanFlow(driver)
    onetimeOTCSkipPayment.OnetimeOTCSkipPayment()


@pytest.fixture()
def OnetimeCompoundSkipPayment(setup):
    onetimeCompoundSkipPayment = CommanFlow(driver)
    onetimeCompoundSkipPayment.OnetimeCompoundSkipPayment()


@pytest.fixture()
def SubscriptionRXSkipPayment(setup):
    SubscriptionRXSkipPayment = CommanFlow(driver)
    SubscriptionRXSkipPayment.SubscriptionRXSkipPayment()


@pytest.fixture()
def SubscriptionOTCSkipPayment(setup):
    SubscriptionOTCSkipPayment = CommanFlow(driver)
    SubscriptionOTCSkipPayment.SubscriptionOTCSkipPayment()


@pytest.fixture()
def SubscriptionCompoundSkipPayment(setup):
    SubscriptionCompoundSkipPayment = CommanFlow(driver)
    SubscriptionCompoundSkipPayment.SubscriptionCompoundSkipPayment()


@pytest.fixture()
def OnetimeRXProvidePayment(setup):
    OnetimeRXProvidePayment = CommanFlow(driver)
    OnetimeRXProvidePayment.OnetimeRXProvidePayment()


@pytest.fixture()
def OnetimeOTCProvidePayment(setup):
    OnetimeOTCProvidePayment = CommanFlow(driver)
    OnetimeOTCProvidePayment.OnetimeOTCProvidePayment()


@pytest.fixture()
def OnetimeCompoundProvidePayment(setup):
    OnetimeCompoundProvidePayment = CommanFlow(driver)
    OnetimeCompoundProvidePayment.OnetimeCompoundProvidePayment()


@pytest.fixture()
def SubscriptionRXProvidePayment(setup):
    SubscriptionRXProvidePayment = CommanFlow(driver)
    SubscriptionRXProvidePayment.SubscriptionRXProvidePayment()


@pytest.fixture()
def SubscriptionOTCProvidePayment(setup):
    SubscriptionOTCProvidePayment = CommanFlow(driver)
    SubscriptionOTCProvidePayment.SubscriptionOTCProvidePayment()


@pytest.fixture()
def SubscriptionCompoundProvidePayment(setup):
    SubscriptionCompoundProvidePayment = CommanFlow(driver)
    SubscriptionCompoundProvidePayment.SubscriptionCompoundProvidePayment()


@pytest.fixture()
def EditOrder(setup):
    editOrder = CommanFlow(driver)
    editOrder.EditOrder()


@pytest.fixture()
def EditPatientDetails(setup):
    editOrder = CommanFlow(driver)
    editOrder.EditPatientDetails()


@pytest.fixture()
def CompleteOrder(setup):
    editOrder = CommanFlow(driver)
    editOrder.CompleteOrder()


@pytest.fixture()
def CancelOrder(setup):
    editOrder = CommanFlow(driver)
    editOrder.CancelOrder()


@pytest.fixture()
def SendOutOfNetwork(setup):
    editOrder = CommanFlow(driver)
    editOrder.SendOutOfNetwork()


@pytest.fixture()
def ApproveOrderFromPractice(setup):
    editOrder = CommanFlow(driver)
    editOrder.ApproveOrderFromPractice()


@pytest.fixture()
def PatientApprovalAndTransferOrder(setup):
    editOrder = CommanFlow(driver)
    editOrder.PatientApprovalAndTransfer()


@pytest.fixture()
def ProcessPaymentAndCreateLabel(setup):
    editOrder = CommanFlow(driver)
    editOrder.ProcessPaymentAndCreateLable()


@pytest.fixture()
def ProcessPaymentAndConfirmPickUpPerson(setup):
    editOrder = CommanFlow(driver)
    editOrder.ProcessPaymentAndConfirmPickUpPerson()


@pytest.fixture()
def ProcessPaymentAndCourierPickUp(setup):
    editOrder = CommanFlow(driver)
    editOrder.ProcessPaymentAndConfirmCourierPickUp()


@pytest.fixture()
def VerifyOrderDetailsScreen(setup):
    orderDetails = CommanFlow(driver)
    orderDetails.VerifyOrderDetailsScreen()


@pytest.mark.hookwrapper
def pytest_runtest_makereport(item):
    """
        Extends the PyTest Plugin to take and embed screenshot in html report, whenever test fails.
        :param item:
        """
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield
    report = outcome.get_result()
    extra = getattr(report, 'extra', [])

    if report.when == 'call' or report.when == "setup":
        xfail = hasattr(report, 'wasxfail')
        if (report.skipped and xfail) or (report.failed and not xfail):
            file_name = report.nodeid.replace("::", "_") + ".png"
            _capture_screenshot(file_name)
            if file_name:
                html = '<div><img src="%s" alt="screenshot" style="width:304px;height:228px;" ' \
                       'onclick="window.open(this.src)" align="right"/></div>' % file_name
                extra.append(pytest_html.extras.html(html))
        report.extra = extra


def _capture_screenshot(name):
    driver.get_screenshot_as_file(name)
