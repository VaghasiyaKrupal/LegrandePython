import time
from Locators.Locators import Locators
from openpyxl import load_workbook

from Locators.PracticeLocators import PracticeLocators
FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']


class TestHub:
    def test_ForgotPassword(self, setup, HubForgotPassword):
        self.driver = setup
        self.driver.close()

    def test_CreatePatient(self, setup, HubLogin, CreatePatient):
        self.driver = setup

    def test_CreateUser(self, setup, HubLogin, CreateUser):
        self.driver = setup

    def test_PatientMainSearch(self, setup, HubLogin, PatientMainSearch):
        self.driver = setup

    def test_ChangeProductPrice(self, setup, HubLogin, ChangeProductPrice):
        self.driver = setup

    def test_CreatingRXOnetimeSkipPayment(self, setup, HubLogin, OnetimeRXSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_EditPatientDetails(self, setup, HubLogin):
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.find_element_by_xpath(Locators.editButton).click()
        self.driver.find_element_by_name(PracticeLocators.cardName).send_keys(testData.cell(2, 7).value)
        self.driver.find_element_by_name(PracticeLocators.maskCardNumberField).send_keys()
        self.driver.find_element_by_name(PracticeLocators.cardCVV).send_keys(testData.cell(2, 9).value)
        self.driver.find_element_by_name(PracticeLocators.addressLine1).send_keys(testData.cell(2, 12).value)
        self.driver.find_element_by_name(PracticeLocators.addressLine2).send_keys(testData.cell(2, 13).value)
        self.driver.find_element_by_name(PracticeLocators.addressCity).send_keys(testData.cell(2, 14).value)
        self.driver.find_element_by_xpath(PracticeLocators.addressState).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        self.driver.find_element_by_name(PracticeLocators.addressZipCode).send_keys(testData.cell(2, 16).value)

    def test_PatientApprovalAndTransferOrder(self, setup, HubLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.find_element_by_xpath(Locators.confirmApprovalButton).click()
        time.sleep(1)
        assert "Patient payment approval has been sucessfully captured."
        dissmissButton = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", dissmissButton)
        self.driver.find_element_by_xpath(Locators.transferOrderButton).click()
        time.sleep(1)
        assert "Have you attached the proper document(s) for this order?"
        self.driver.find_element_by_xpath(Locators.yesButton).click()
        time.sleep(1)
        assert "This order has been successfully transferred."
        time.sleep(2)
        Button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", Button)
        self.driver.quit()

    def test_CompleteOrder(self, setup, HubLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.find_element_by_xpath(Locators.selectActionDropdown).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(Locators.completeValue).click()
        self.driver.find_element_by_class_name('form-control').send_keys('Patient completed the treatment.')
        # self.driver.find_element_by_xpath(PracticeLocators.submitButton).click()
        time.sleep(2)
        assert "You have successfully completed this order"
        button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", button)
        self.driver.quit()

    def test_CreatingOTCOnetimeSkipPayment(self, setup, HubLogin, OnetimeOTCSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_OrderSendOutOfNetwork(self, setup, HubLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.find_element_by_xpath(Locators.selectActionDropdown).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(Locators.sendOutOfNetwork).click()
        self.driver.find_element_by_class_name('form-control').send_keys('Order send out of network.')
        # self.driver.find_element_by_xpath(PracticeLocators.submitButton).click()
        time.sleep(2)
        assert "You have successfully sent this order out of network"
        button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", button)
        self.driver.quit()

    def test_CreatingCompoundOnetimeSkipPayment(self, setup, HubLogin, OnetimeCompoundSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_CancelOrder(self, setup, HubLogin):
        self.driver = setup
        self.driver.find_element_by_css_selector(Locators.patientSearch).send_keys(testData.cell(2, 1).value)
        time.sleep(3)
        self.driver.find_element_by_css_selector(Locators.firstOrder).click()
        self.driver.find_element_by_xpath(Locators.selectActionDropdown).click()
        time.sleep(2)
        self.driver.find_element_by_xpath(Locators.cancelValue).click()
        self.driver.find_element_by_class_name('form-control').send_keys('I do not need this order.')
        self.driver.find_element_by_xpath(PracticeLocators.submitButton).click()
        time.sleep(2)
        assert "You have successfully canceled this order"
        button = self.driver.find_element_by_xpath(PracticeLocators.dismissButton)
        self.driver.execute_script("arguments[0].click()", button)
        self.driver.quit()

    def test_CreatingRXSubscriptionSkipPayment(self, setup, HubLogin, SubscriptionRXSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_CreatingOTCSubscriptionSkipPayment(self, setup, HubLogin, SubscriptionOTCSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_CreatingCompoundSubscriptionSkipPayment(self, setup, HubLogin, SubscriptionCompoundSkipPayment):
        self.driver = setup
        self.driver.close()

    def test_EditOrder(self, setup, HubLogin, EditOrder):
        self.driver = setup

    def test_RXOnetimeProvidePayment(self, setup, HubLogin, OnetimeRXProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_OTCOnetimeProvidePayment(self, setup, HubLogin, OnetimeOTCProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_CompoundOnetimeProvidePayment(self, setup, HubLogin, OnetimeCompoundProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_RXSubscriptionProvidePayment(self, setup, HubLogin, SubscriptionRXProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_OTCSubscriptionProvidePayment(self, setup, HubLogin, SubscriptionOTCProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_CompoundSubscriptionProvidePayment(self, setup, HubLogin, SubscriptionCompoundProvidePayment):
        self.driver = setup
        self.driver.close()

    def test_EditOrder(self, setup, HubLogin, EditOrder):
        self.driver = setup
