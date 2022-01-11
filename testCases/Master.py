import time
import random
from faker import Faker
from datetime import date
from openpyxl import load_workbook
from Locators.Locators import Locators
from selenium.webdriver.common.by import By
from Locators.MasterLocators import MasterLocators
from Locators.PracticeLocators import PracticeLocators
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
scriptData = datafile['Script Data']
loginData = datafile['Login Credentials']

faker = Faker()
today = date.today()
currentDate = today.strftime("%m/%d/%Y")
firstName = faker.first_name()
lastName = faker.last_name()
street_1 = faker.address()
street_2 = faker.address()
city = faker.city()
zipCode = faker.zipcode()
PhoneNumber = faker.phone_number()
categoryName = "Automation"

rxProductName = "Selenium " + firstName + " RX"
otcProductName = "Selenium " + firstName + " OTC"
compoundProductName = "Selenium " + firstName + " Compound"

RxDispenserName = "Selenium " + firstName + " RX Dispensers"
otcDispenserName = "Selenium " + firstName + " OTC Dispensers"
CompoundDispenserName = "Selenium " + firstName + " Compound Dispenser"
hubDispenserName = "Selenium " + firstName + " Hub Dispensers"

rxDispenserEmail = firstName.lower() + ".rx"
otcDispenserEmail = firstName.lower() + ".otc"
compoundDispenserEmail = firstName.lower() + ".compound"
hubDispenserEmail = firstName.lower() + ".hub"
practiceEmail = firstName.lower() + ".practice"
emailAddress = "@mailinator.com"
UserEmailPractice = ".practice@mailinator.com"


class TestForgotPassword:
    def test_ForgotPassword(self, setup, MasterForgotPassword):
        self.driver = setup
        self.driver.quit()


class TestProductCreation:
    def test_CreateRXProduct(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.addProductLink).click()
        assert self.driver.find_element_by_xpath(MasterLocators.activeProductCheckbox).is_selected()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys(rxProductName)
        scriptData.cell(2, 1).value = rxProductName
        datafile.save(FilePath)
        self.driver.find_element_by_class_name(MasterLocators.selectArrow).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.rxProductType).click()
        self.driver.find_element_by_name(MasterLocators.categoryField).send_keys(categoryName)
        self.driver.find_element_by_name(MasterLocators.brandField).send_keys(categoryName)
        self.driver.find_element_by_name(MasterLocators.packageField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.priceField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.gmvField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.descriptionField).send_keys(MasterLocators.description)
        self.driver.find_element_by_name(MasterLocators.awpField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.arpField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.ndcField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.pnField).send_keys(random.randint(0, 100))
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.uploadImage).send_keys('C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Image.png')
        time.sleep(5)
        self.driver.find_element_by_xpath(MasterLocators.addProductButton).click()
        time.sleep(2)
        assert "You have successfully added "+scriptData.cell(2, 1).value + "."
        self.driver.quit()

    def test_CreateOTCProduct(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.addProductLink).click()
        assert self.driver.find_element_by_xpath(MasterLocators.activeProductCheckbox).is_selected()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys(otcProductName)
        scriptData.cell(2, 2).value = otcProductName
        datafile.save(FilePath)
        self.driver.find_element_by_class_name(MasterLocators.selectArrow).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.otcProductType).click()
        self.driver.find_element_by_name(MasterLocators.categoryField).send_keys(categoryName)
        self.driver.find_element_by_name(MasterLocators.brandField).send_keys(categoryName)
        self.driver.find_element_by_name(MasterLocators.packageField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.priceField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.gmvField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.descriptionField).send_keys(MasterLocators.description)
        self.driver.find_element_by_name(MasterLocators.awpField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.arpField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.ndcField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.pnField).send_keys(random.randint(0, 100))
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.uploadImage).send_keys('C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Image.png')
        time.sleep(5)
        self.driver.find_element_by_xpath(MasterLocators.addProductButton).click()
        time.sleep(2)
        assert "You have successfully added "+scriptData.cell(2, 2).value + "."
        self.driver.quit()

    def test_CreateCompoundProduct(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.addProductLink).click()
        assert self.driver.find_element_by_xpath(MasterLocators.activeProductCheckbox).is_selected()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys(compoundProductName)
        scriptData.cell(2, 3).value = compoundProductName
        datafile.save(FilePath)
        self.driver.find_element_by_class_name(MasterLocators.selectArrow).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.compoundProductType).click()
        self.driver.find_element_by_name(MasterLocators.categoryField).send_keys(categoryName)
        self.driver.find_element_by_name(MasterLocators.brandField).send_keys(categoryName)
        self.driver.find_element_by_name(MasterLocators.packageField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.priceField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.gmvField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.descriptionField).send_keys(MasterLocators.description)
        self.driver.find_element_by_name(MasterLocators.awpField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.arpField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.ndcField).send_keys(random.randint(0, 100))
        self.driver.find_element_by_name(MasterLocators.pnField).send_keys(random.randint(0, 100))
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.uploadImage).send_keys('C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Image.png')
        time.sleep(5)
        self.driver.find_element_by_xpath(MasterLocators.addProductButton).click()
        time.sleep(2)
        assert "You have successfully added "+scriptData.cell(2, 3).value + "."
        self.driver.quit()


class TestDispenserCreation:
    # def test_CreateHubDispenser(self, setup, MasterLogin):
    #     self.driver = setup
    #     self.driver.find_element_by_link_text(MasterLocators.dispenserScreen).click()
    #     self.driver.find_element_by_link_text(MasterLocators.addDispenserLink).click()
    #     self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " Hub")
    #     scriptData.cell(2, 4).value = "Selenium " + firstName + " Hub"
    #     loginData.cell(4, 3).value = hubDispenserEmail + emailAddress
    #     datafile.save(FilePath)
    #     self.driver.find_element_by_name(MasterLocators.emailField).send_keys(loginData.cell(4, 3).value)
    #     self.driver.find_element_by_name(MasterLocators.phoneField).send_keys(PhoneNumber)
    #     self.driver.find_element_by_name(MasterLocators.faxField).send_keys(PhoneNumber)
    #     self.driver.find_element_by_name(MasterLocators.addressStreet1Field).send_keys(street_1)
    #     self.driver.find_element_by_name(MasterLocators.addressCityField).send_keys(city)
    #     self.driver.find_element_by_xpath(MasterLocators.selectState).click()
    #     time.sleep(1)
    #     self.driver.find_element_by_xpath(PracticeLocators.state).click()
    #     self.driver.find_element_by_name(MasterLocators.addressZipCodeField).send_keys(zipCode)
    #     self.driver.find_element_by_xpath(MasterLocators.selectTimezone).click()
    #     time.sleep(1)
    #     self.driver.find_element_by_xpath(MasterLocators.timeZoneValue).click()
    #     self.driver.find_element_by_xpath(MasterLocators.selectHubCategory).click()
    #     self.driver.find_element_by_name(MasterLocators.npiField).send_keys(random.randint(0, 9999999))
    #     self.driver.find_element_by_name(MasterLocators.licenseField).send_keys(random.randint(0, 9999999))
    #     self.driver.find_element_by_name(MasterLocators.adminFirstNameField).send_keys(firstName)
    #     self.driver.find_element_by_name(MasterLocators.adminLastnameField).send_keys(lastName)
    #     self.driver.find_element_by_name(MasterLocators.adminEmailField).send_keys(loginData.cell(4, 3).value)
    #     self.driver.find_element_by_name(MasterLocators.adminPhoneField).send_keys(PhoneNumber)
    #     self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
    #     time.sleep(2)
    #     assert "Congratulations. You have successfully created a new dispenser account."
    #     self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
    #     self.driver.quit()
    #
    # def test_CreateHubEmailAndSetPassword(self, setup):
    #     self.driver = setup
    #     self.driver.get(Locators.malinatorLink)
    #     self.driver.implicitly_wait(10)
    #     self.driver.find_element_by_id(Locators.searchBoxID).send_keys(loginData.cell(4, 3).value)
    #     self.driver.find_element_by_xpath(Locators.goButtonLink).click()
    #     self.driver.implicitly_wait(10)
    #     self.driver.find_element_by_xpath(MasterLocators.emailInbox).click()
    #     time.sleep(3)
    #     WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, Locators.emailIframe)))
    #     WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, MasterLocators.accountSetupLink))).click()
    #     self.driver.switch_to.window(self.driver.window_handles[1])
    #     self.driver.implicitly_wait(5)
    #     self.driver.find_element_by_name(MasterLocators.newPassword).send_keys(loginData.cell(3, 4).value)
    #     self.driver.find_element_by_name(MasterLocators.confirmPassword).send_keys(loginData.cell(3, 4).value)
    #     time.sleep(1)
    #     self.driver.find_element_by_xpath(MasterLocators.setPasswordButton).click()
    #     time.sleep(2)
    #     assert "Password set successfully"
    #     self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
    #     self.driver.switch_to.window(self.driver.window_handles[0])
    #     self.driver.quit()

    def test_CreateRXDispenser(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.dispenserScreen).click()
        self.driver.find_element_by_link_text(MasterLocators.addDispenserLink).click()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " RX")
        scriptData.cell(2, 5).value = "Selenium " + firstName + " RX"
        loginData.cell(5, 3).value = rxDispenserEmail+emailAddress
        datafile.save(FilePath)
        self.driver.find_element_by_name(MasterLocators.emailField).send_keys(loginData.cell(5, 3).value)
        self.driver.find_element_by_name(MasterLocators.phoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.faxField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.addressStreet1Field).send_keys(street_1)
        self.driver.find_element_by_name(MasterLocators.addressCityField).send_keys(city)
        self.driver.find_element_by_xpath(MasterLocators.selectState).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        self.driver.find_element_by_name(MasterLocators.addressZipCodeField).send_keys(zipCode)
        self.driver.find_element_by_xpath(MasterLocators.selectTimezone).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.timeZoneValue).click()
        self.driver.find_element_by_name(MasterLocators.npiField).send_keys(random.randint(0, 9999999))
        self.driver.find_element_by_name(MasterLocators.licenseField).send_keys(random.randint(0, 9999999))
        self.driver.find_element_by_name(MasterLocators.adminFirstNameField).send_keys(firstName)
        self.driver.find_element_by_name(MasterLocators.adminLastnameField).send_keys(lastName)
        self.driver.find_element_by_name(MasterLocators.adminEmailField).send_keys(loginData.cell(5, 3).value)
        self.driver.find_element_by_name(MasterLocators.adminPhoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
        time.sleep(2)
        assert "Congratulations. You have successfully created a new dispenser account."
        self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
        self.driver.quit()

    def test_CreateRXEmailAndSetPassword(self, setup):
        self.driver = setup
        self.driver.get(Locators.malinatorLink)
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_id(Locators.searchBoxID).send_keys(loginData.cell(5, 3).value)
        self.driver.find_element_by_xpath(Locators.goButtonLink).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(MasterLocators.emailInbox).click()
        time.sleep(3)
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, Locators.emailIframe)))
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, MasterLocators.accountSetupLink))).click()
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.implicitly_wait(5)
        self.driver.find_element_by_name(MasterLocators.newPassword).send_keys(loginData.cell(3, 4).value)
        self.driver.find_element_by_name(MasterLocators.confirmPassword).send_keys(loginData.cell(3, 4).value)
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.setPasswordButton).click()
        time.sleep(2)
        assert "Password set successfully"
        self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.quit()

    def test_CreateOTCDispenser(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.dispenserScreen).click()
        self.driver.find_element_by_link_text(MasterLocators.addDispenserLink).click()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " OTC")
        scriptData.cell(2, 6).value = "Selenium " + firstName + " OTC"
        loginData.cell(6, 3).value = otcDispenserEmail + emailAddress
        datafile.save(FilePath)
        self.driver.find_element_by_name(MasterLocators.emailField).send_keys(loginData.cell(6, 3).value)
        self.driver.find_element_by_name(MasterLocators.phoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.faxField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.addressStreet1Field).send_keys(street_1)
        self.driver.find_element_by_name(MasterLocators.addressCityField).send_keys(city)
        self.driver.find_element_by_xpath(MasterLocators.selectState).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        self.driver.find_element_by_name(MasterLocators.addressZipCodeField).send_keys(zipCode)
        self.driver.find_element_by_xpath(MasterLocators.selectTimezone).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.timeZoneValue).click()
        self.driver.find_element_by_name(MasterLocators.npiField).send_keys(random.randint(0, 9999999))
        self.driver.find_element_by_name(MasterLocators.licenseField).send_keys(random.randint(0, 9999999))
        self.driver.find_element_by_name(MasterLocators.adminFirstNameField).send_keys(firstName)
        self.driver.find_element_by_name(MasterLocators.adminLastnameField).send_keys(lastName)
        self.driver.find_element_by_name(MasterLocators.adminEmailField).send_keys(loginData.cell(6, 3).value)
        self.driver.find_element_by_name(MasterLocators.adminPhoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
        time.sleep(2)
        assert "Congratulations. You have successfully created a new dispenser account."
        self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
        self.driver.quit()

    def test_CreateOTCEmailAndSetPassword(self, setup):
        self.driver = setup
        self.driver.get(Locators.malinatorLink)
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_id(Locators.searchBoxID).send_keys(loginData.cell(6, 3).value)
        self.driver.find_element_by_xpath(Locators.goButtonLink).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(MasterLocators.emailInbox).click()
        time.sleep(3)
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, Locators.emailIframe)))
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, MasterLocators.accountSetupLink))).click()
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.implicitly_wait(5)
        self.driver.find_element_by_name(MasterLocators.newPassword).send_keys(loginData.cell(3, 4).value)
        self.driver.find_element_by_name(MasterLocators.confirmPassword).send_keys(loginData.cell(3, 4).value)
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.setPasswordButton).click()
        time.sleep(2)
        assert "Password set successfully"
        self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.quit()

    def test_CreateCompoundDispenser(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.dispenserScreen).click()
        self.driver.find_element_by_link_text(MasterLocators.addDispenserLink).click()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " Compound")
        scriptData.cell(2, 7).value = "Selenium " + firstName + " Compound"
        loginData.cell(7, 3).value = compoundDispenserEmail + emailAddress
        datafile.save(FilePath)
        self.driver.find_element_by_name(MasterLocators.emailField).send_keys(loginData.cell(7, 3).value)
        self.driver.find_element_by_name(MasterLocators.phoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.faxField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.addressStreet1Field).send_keys(street_1)
        self.driver.find_element_by_name(MasterLocators.addressCityField).send_keys(city)
        self.driver.find_element_by_xpath(MasterLocators.selectState).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        self.driver.find_element_by_name(MasterLocators.addressZipCodeField).send_keys(zipCode)
        self.driver.find_element_by_xpath(MasterLocators.selectTimezone).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.timeZoneValue).click()
        self.driver.find_element_by_name(MasterLocators.npiField).send_keys(random.randint(0, 9999999))
        self.driver.find_element_by_name(MasterLocators.licenseField).send_keys(random.randint(0, 9999999))
        self.driver.find_element_by_name(MasterLocators.adminFirstNameField).send_keys(firstName)
        self.driver.find_element_by_name(MasterLocators.adminLastnameField).send_keys(lastName)
        self.driver.find_element_by_name(MasterLocators.adminEmailField).send_keys(loginData.cell(7, 3).value)
        self.driver.find_element_by_name(MasterLocators.adminPhoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
        time.sleep(2)
        assert "Congratulations. You have successfully created a new dispenser account."
        self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
        self.driver.quit()

    def test_CreateCompoundEmailAndSetPassword(self, setup):
        self.driver = setup
        self.driver.get(Locators.malinatorLink)
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_id(Locators.searchBoxID).send_keys(loginData.cell(7, 3).value)
        self.driver.find_element_by_xpath(Locators.goButtonLink).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(MasterLocators.emailInbox).click()
        time.sleep(3)
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, Locators.emailIframe)))
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, MasterLocators.accountSetupLink))).click()
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.implicitly_wait(5)
        self.driver.find_element_by_name(MasterLocators.newPassword).send_keys(loginData.cell(3, 4).value)
        self.driver.find_element_by_name(MasterLocators.confirmPassword).send_keys(loginData.cell(3, 4).value)
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.setPasswordButton).click()
        time.sleep(2)
        assert "Password set successfully"
        self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.quit()


class TestPracticeCreation:
    def test_CreatePracticeAccount(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.practiceScreen).click()
        self.driver.find_element_by_link_text(MasterLocators.addPracticeLink).click()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " Practice")
        scriptData.cell(2, 8).value = "Selenium " + firstName
        loginData.cell(3, 3).value = practiceEmail + emailAddress
        datafile.save(FilePath)
        self.driver.find_element_by_name(MasterLocators.emailField).send_keys(loginData.cell(3, 3).value)
        self.driver.find_element_by_name(MasterLocators.phoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.faxField).send_keys(PhoneNumber)
        self.driver.find_element_by_name(MasterLocators.addressStreet1Field).send_keys(street_1)
        self.driver.find_element_by_name(MasterLocators.addressCityField).send_keys(city)
        self.driver.find_element_by_xpath(MasterLocators.selectState).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.state).click()
        self.driver.find_element_by_name(MasterLocators.addressZipCodeField).send_keys(zipCode)
        self.driver.find_element_by_xpath(MasterLocators.selectTimezone).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.timeZoneValue).click()
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.practiceValue).click()
        self.driver.find_element_by_name(MasterLocators.npiField).send_keys(random.randint(0, 9999999))
        self.driver.find_element_by_name(MasterLocators.licenseField).send_keys(random.randint(0, 9999999))

        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.RXDispenser).click()
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.OTCDispenser).click()
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.CompoundDispenser).click()

        self.driver.find_element_by_name(MasterLocators.adminFirstNameField).send_keys("Selenium " + firstName)
        self.driver.find_element_by_name(MasterLocators.adminLastnameField).send_keys("Practice")
        self.driver.find_element_by_name(MasterLocators.adminEmailField).send_keys(loginData.cell(3, 3).value)
        self.driver.find_element_by_name(MasterLocators.adminPhoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
        time.sleep(2)
        assert "Congratulations. You have successfully created a new dispenser account."
        self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
        self.driver.quit()

    def test_CreatePracticeEmailAndSetPassword(self, setup):
        self.driver = setup
        self.driver.get(Locators.malinatorLink)
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_id(Locators.searchBoxID).send_keys(loginData.cell(3, 3).value)
        self.driver.find_element_by_xpath(Locators.goButtonLink).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(MasterLocators.emailInbox).click()
        time.sleep(3)
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, Locators.emailIframe)))
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, MasterLocators.accountSetupLink))).click()
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.implicitly_wait(5)
        self.driver.find_element_by_name(MasterLocators.newPassword).send_keys(loginData.cell(3, 4).value)
        self.driver.find_element_by_name(MasterLocators.confirmPassword).send_keys(loginData.cell(3, 4).value)
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.setPasswordButton).click()
        time.sleep(2)
        assert "Password set successfully"
        self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.quit()


class TestUserCreation:
    def test_CreateUserForPractice(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.userScreen).click()
        self.driver.find_element_by_xpath(PracticeLocators.addUserButton).click()
        self.driver.find_element_by_name(PracticeLocators.firstName).send_keys(firstName)
        self.driver.find_element_by_name(PracticeLocators.lastName).send_keys(lastName)
        scriptData.cell(2, 9).value = firstName + " " + lastName
        loginData.cell(9, 3).value = firstName + "." + lastName + UserEmailPractice
        datafile.save(FilePath)
        self.driver.find_element_by_name(MasterLocators.emailField).send_keys(loginData.cell(9, 3).value)
        self.driver.find_element_by_name(MasterLocators.phoneField).send_keys(PhoneNumber)
        self.driver.find_element_by_xpath(MasterLocators.practiceRadio).click()
        self.driver.find_element_by_xpath(MasterLocators.searchPractice).send_keys(scriptData.cell(2, 8).value)
        time.sleep(2)
        self.driver.find_element_by_xpath(MasterLocators.selectUserPractice).click()
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.userType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(PracticeLocators.createAccountButton).click()
        time.sleep(3)
        assert "Congratulations. You have successfully created new user profile."
        self.driver.quit()

    def test_CreateUserEmailAndSetPassword(self, setup):
        self.driver = setup
        self.driver.get(Locators.malinatorLink)
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_id(Locators.searchBoxID).send_keys(loginData.cell(9, 3).value)
        self.driver.find_element_by_xpath(Locators.goButtonLink).click()
        self.driver.implicitly_wait(10)
        self.driver.find_element_by_xpath(MasterLocators.emailInbox).click()
        time.sleep(3)
        WebDriverWait(self.driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH, Locators.emailIframe)))
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, MasterLocators.accountSetupLink))).click()
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.implicitly_wait(5)
        self.driver.find_element_by_name(MasterLocators.newPassword).send_keys(loginData.cell(3, 4).value)
        self.driver.find_element_by_name(MasterLocators.confirmPassword).send_keys(loginData.cell(3, 4).value)
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.setPasswordButton).click()
        time.sleep(2)
        assert "Password set successfully"
        self.driver.find_element_by_class_name(MasterLocators.closeButton).click()
        self.driver.switch_to.window(self.driver.window_handles[0])
        self.driver.quit()


class TestPayorCreation:
    def test_CreateInsurancePayor(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.PayorScreen).click()
        self.driver.find_element_by_xpath(MasterLocators.addPayorButton).click()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " Insurance")
        scriptData.cell(2, 10).value = "Selenium " + firstName + " Insurance"
        datafile.save(FilePath)
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.insurance).click()
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.payorStatus).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.createPayorButton).click()
        time.sleep(3)
        assert "Payor has been successfully created"
        self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
        self.driver.quit()

    def test_CreateCashPayor(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.PayorScreen).click()
        self.driver.find_element_by_xpath(MasterLocators.addPayorButton).click()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " Cash")
        scriptData.cell(2, 11).value = "Selenium " + firstName + " Cash"
        datafile.save(FilePath)
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.Cash).click()
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.payorStatus).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.createPayorButton).click()
        time.sleep(3)
        assert "Payor has been successfully created"
        self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
        self.driver.quit()

    def test_CreateCouponPayor(self, setup, MasterLogin):
        self.driver = setup
        self.driver.find_element_by_link_text(MasterLocators.PayorScreen).click()
        self.driver.find_element_by_xpath(MasterLocators.addPayorButton).click()
        self.driver.find_element_by_name(MasterLocators.nameField).send_keys("Selenium " + firstName + " Coupon")
        scriptData.cell(2, 12).value = "Selenium " + firstName + " Coupon"
        datafile.save(FilePath)
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.Coupon).click()
        self.driver.find_element_by_xpath(MasterLocators.dropdownSelectType).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.payorStatus).click()
        time.sleep(1)
        self.driver.find_element_by_xpath(MasterLocators.createPayorButton).click()
        time.sleep(3)
        assert "Payor has been successfully created"
        self.driver.find_element_by_xpath(MasterLocators.dismissButton).click()
        self.driver.quit()
