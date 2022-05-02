import time
import os
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from Locators.PatientLocators import PatientLocators
from Locators.PracticeLocators import PracticeLocators
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

FilePath = "TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']
scriptData = datafile['Script Data']


class TestPractice:
    # def test_ForgotPassword(self, setup, PracticeForgotPassword):
    #     self.driver = setup
    #     self.driver.quit()
    #
    # def test_CreatePatient(self, setup, PracticeLogin, CreatePatient):
    #     self.driver = setup
    #     self.driver.quit()
    #
    # def test_onetimeTemplate(self, setup, PracticeLogin):
    #     self.driver = setup
    #     WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, PracticeLocators.orderTemplateScreen))).click()
    #     self.driver.find_element_by_link_text(PracticeLocators.createTemplateButton).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.selectOnetimeTemplate).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.templateTitle).send_keys(testData.cell(2, 6).value)
    #     self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.templateProductSearch).send_keys(scriptData.cell(2, 1).value)
    #     time.sleep(2)
    #     self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 1).value + "']").click()
    #     self.driver.find_element_by_xpath(PracticeLocators.AddButton).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
    #     time.sleep(1)
    #     self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
    #     time.sleep(1)
    #     self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
    #     time.sleep(2)
    #     self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
    #     time.sleep(2)
    #     instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
    #     instruction.click()
    #     instruction.send_keys(testData.cell(2, 3).value)
    #     self.driver.find_element_by_xpath(PracticeLocators.templateNotes).send_keys(testData.cell(2, 5).value)
    #     self.driver.find_element_by_xpath(PracticeLocators.saveTemplateButton).click()
    #     self.driver.quit()
    #
    # def test_subscriptionTemplate(self, setup, PracticeLogin):
    #     self.driver = setup
    #     WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, PracticeLocators.orderTemplateScreen))).click()
    #     self.driver.find_element_by_link_text(PracticeLocators.createTemplateButton).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.selectSubscriptionTemplate).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.templateTitle).send_keys(testData.cell(2, 6).value)
    #     self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.templateProductSearch).send_keys(scriptData.cell(2, 1).value)
    #     time.sleep(2)
    #     self.driver.find_element_by_xpath("//*[text()='" + scriptData.cell(2, 1).value + "']").click()
    #     self.driver.find_element_by_xpath(PracticeLocators.addProductButton).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.ProductQuantity).click()
    #     time.sleep(1)
    #     self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
    #     self.driver.find_element_by_xpath(PracticeLocators.ProductRefilles).click()
    #     time.sleep(1)
    #     self.driver.find_element_by_xpath(PracticeLocators.Quantity).click()
    #     time.sleep(2)
    #     self.driver.find_element_by_xpath(PracticeLocators.DAWCheckbox).click()
    #     time.sleep(2)
    #     instruction = self.driver.find_element_by_xpath(PracticeLocators.productInstruction)
    #     instruction.click()
    #     instruction.send_keys(testData.cell(2, 3).value)
    #     self.driver.find_element_by_xpath(PracticeLocators.templateNotes).send_keys(testData.cell(2, 5).value)
    #     self.driver.find_element_by_xpath(PracticeLocators.saveTemplateButton).click()
    #     self.driver.quit()

    def test_DropChart(self, setup, PracticeLogin):
        self.driver = setup
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, PracticeLocators.documentsScreen))).click()
        self.driver.find_element_by_xpath(PracticeLocators.newDocuments).click()
        self.driver.find_element_by_xpath(PracticeLocators.documentsDoctorSearchbox).click()
        self.driver.find_element_by_xpath(PracticeLocators.selectDoctor).click()
        self.driver.find_element_by_xpath(PracticeLocators.selectDocumentsSearchbox).click()
        self.driver.find_element_by_xpath(PracticeLocators.documentsType).click()
        self.driver.find_element_by_xpath(PracticeLocators.documentTitle).send_keys('WL')
        self.driver.find_element_by_xpath(PracticeLocators.continueButton).click()
        self.driver.find_element_by_css_selector(PracticeLocators.uploadFile).send_keys(os.path.abspath("TestData/WelcomeLetter.pdf"))
        SubmitButton = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, PracticeLocators.submitButton)))
        if SubmitButton.is_enabled():
            SubmitButton.click()
        else:
            print("\n File not uploaded")
        self.driver.quit()

    # def test_createUser(self, setup, PracticeLogin, CreateUser):
    #     self.driver = setup
    #     self.driver.quit()
    #
    # def test_PatientMainSearch(self, setup, PracticeLogin, PatientMainSearch):
    #     self.driver = setup
    #     self.driver.quit()
    #
    # def test_OnetimeRXSkipPayment(self, setup, PracticeLogin, OnetimeRXSkipPayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Onetime RX with skip payment, created successfully')
    #     else:
    #         print("\nOrder: Onetime RX with skip payment, is not created")
    #     self.driver.quit()
    #
    # def test_VerifyOrderDetailsScreen(self, setup, PracticeLogin, VerifyOrderDetailsScreen):
    #     self.driver = setup
    #     self.driver.quit()
    #
    # def test_OnetimeOTCSkipPayment(self, setup, PracticeLogin, OnetimeOTCSkipPayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Onetime OTC with skip payment, created successfully')
    #     else:
    #         print("\nOrder: Onetime OTC with skip payment, is not created")
    #     self.driver.quit()
    #
    # def test_OnetimeCompoundSkipPayment(self, setup, PracticeLogin, OnetimeCompoundSkipPayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Onetime Compound with skip payment, created successfully')
    #     else:
    #         print("\nOrder: Onetime Compound with skip payment, is not created")
    #     self.driver.quit()
    #
    # def test_SubscriptionRXSkipPayment(self, setup, PracticeLogin, SubscriptionRXSkipPayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Subscription RX with Skip payment, created successfully')
    #     else:
    #         print("\nOrder: Subscription RX with Skip payment, is not created")
    #     self.driver.quit()
    #
    # def test_SubscriptionOTCSkipPayment(self, setup, PracticeLogin, SubscriptionOTCSkipPayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Subscription OTC with Skip payment, created successfully')
    #     else:
    #         print("\nOrder: Subscription OTC with Skip payment, is not created")
    #     self.driver.quit()
    #
    # def test_SubscriptionCompoundSkipPayment(self, setup, PracticeLogin, SubscriptionCompoundSkipPayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Subscription Compound with Skip payment, created successfully')
    #     else:
    #         print("\nOrder: Subscription Compound with Skip payment, is not created")
    #     self.driver.quit()
    #
    # def test_EditOrder(self, setup, PracticeLogin, EditOrder):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nYour order is successfully updated')
    #     else:
    #         print("\nYour order is not update")
    #     self.driver.quit()
    #
    # def test_OnetimeRXProvidePayment(self, setup, PracticeLogin, OnetimeRXProvidePayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Onetime RX with provide payment, created successfully')
    #     else:
    #         print("\nOrder: Onetime RX with provide payment, is not created")
    #     self.driver.quit()
    #
    # def test_OnetimeOTCProvidePayment(self, setup, PracticeLogin, OnetimeOTCProvidePayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Onetime OTC with provide payment, created successfully')
    #     else:
    #         print("\nOrder: Onetime OTC with provide payment, is not created")
    #     self.driver.quit()
    #
    # def test_OnetimeCompoundProvidePayment(self, setup, PracticeLogin, OnetimeCompoundProvidePayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Onetime Compound with provide payment, created successfully')
    #     else:
    #         print("\nOrder: Onetime Compound with provide payment, is not created")
    #     self.driver.quit()
    #
    # def test_SubscriptionRXProvidePayment(self, setup, PracticeLogin, SubscriptionRXProvidePayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Subscription RX with provide payment, created successfully')
    #     else:
    #         print("\nOrder: Subscription RX with provide payment, is not created")
    #     self.driver.quit()
    #
    # def test_SubscriptionOTCProvidePayment(self, setup, PracticeLogin, SubscriptionOTCProvidePayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Subscription OTC with provide payment, created successfully')
    #     else:
    #         print("\nOrder: Subscription OTC with provide payment, is not created")
    #     self.driver.quit()
    #
    # def test_SubscriptionCompoundProvidePayment(self, setup, PracticeLogin, SubscriptionCompoundProvidePayment):
    #     self.driver = setup
    #     if "You're all set!" in self.driver.page_source:
    #         self.driver.find_element_by_partial_link_text("Return to Dashboard").click()
    #         print('\nOrder: Subscription Compound with provide payment, created successfully')
    #     else:
    #         print("\nOrder: Subscription Compound with provide payment, is not created")
    #     self.driver.quit()
    #
    # def test_CreateOrderFromUserAccount(self, setup, UserLogin, OnetimeRXSkipPayment):
    #     self.driver = setup
    #     # number = self.driver.find_element_by_xpath('//div[@class="sc-kafWEX ihwrOP"]//div/div/div[3]').text
    #     # desired_text = number.split(':')[1]
    #     # print("\n", desired_text)
    #     self.driver.quit()
    #
    # def test_CheckUserOrder(self, setup, PracticeLogin):
    #     self.driver = setup
    #     time.sleep(3)
    #     assert self.driver.find_element_by_xpath(PatientLocators.myQueueOrderDate)
    #     assert self.driver.find_element_by_xpath(PatientLocators.myQueuePatientName)
    #     self.driver.quit()
