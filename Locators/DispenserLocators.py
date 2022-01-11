from datetime import date
from openpyxl import load_workbook

FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']

today = date.today()
currentDate = today.strftime("%m/%d/%Y")


class DispenserLocators:
    productList = "Product List"
    updateButton = '//*[text()="Update"]'
    changePriceButton = '//*[text()="Change Item Price"]'
    confirmApprovalButton = '//*[text()="Confirm Approval"]'
    transferOrderButton = '//*[text()="Transfer Order"]'
    yesButton = '//*[text()="Yes"]'
    editButton = '(//*[text()="Edit"])[4]'
    patientSearch = '[placeholder="Search by Order Number, Patient Name, or DOB (mm/dd/yyyy)"]'
    firstOrder = '[data-expanded="false"]'
    addNoteButton = '//*[text()="Add Note"]'
    completeValue = '//*[@aria-label="Complete"]'
    sendOutOfNetwork = '//*[@aria-label="Send out of Network"]'
    cancelValue = '//*[@aria-label="Cancel"]'
    selectActionDropdown = '//div[@class="sc-eTpRJs ihfShO"]//*[@class="Select-arrow"]'
    saveAndExitButton = '//button[text()="Save and Exit"]'
    selectServiceType = '//*[@class="sc-chbbiW eBhBQG"]//span[@class="Select-arrow"]'
    pickUpPerson = '[aria-label="Pick Up In Person"]'
    courierService = '[aria-label="Courier Service"]'
    confirmCourierPickUp = '//button[text()="Confirm Courier Pick-Up"]'
    completeButton = '//button[text()="Complete"]'
    editPaymentButton = '//div[@id="card"]//*[text()="Edit"]'

    # STD Pharma Locators
    approveButton = '//*[text()="Approve"]'
    verifyOrderDate = "(//tr[@class='table-row'])[1]//td[2][text()='" + currentDate + "']"
    verifyPatientName = "(//tr[@class='table-row'])[1]//td[3][text()='" + testData.cell(2, 1).value + "']"
    processPaymentButton = '//button[text()="Process Payment"]'
    createPostageLabelButton = '//div[@class="modal-footer"]//button[text()="Create Postage Label"]'
    createLabelButton = '//div[@class="modal-footer"]//button[text()="Create Label"]'
    printLabelButton = '//div[@class="modal-footer"]//a[text()="Print Postage Label"]'
    confirmPickUpButton = '//button[text()="Confirm Pick-Up"]'
