from openpyxl import load_workbook

FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']
scriptData = datafile['Script Data']


class PracticeLocators:
    addPersonButton = "person_add"
    firstName = "first_name"
    lastName = "last_name"
    birthDate = "date_of_birth"
    phoneNumber = "phone"
    closeButton = "//button[@type='button' and text()='Close']"
    Add_RX = "add_rx"
    orderTemplateScreen = 'Order Templates'
    createTemplateButton = 'New Plan Template'
    documentsScreen = 'Documents'
    newDocuments = '//*[text()="Upload New Document"]'
    documentsDoctorSearchbox = '//*[@placeholder="Search by first or last name"]'
    selectDoctor = "//div[@data-test='_HINTS_HINT']//*[text()='" + scriptData.cell(2, 8).value + " Practice']"
    selectDocumentsSearchbox = '//*[text()="Select a document type"]'
    documentsType = '//*[@class="Select-menu-outer"]//*[text()="Welcome Letter"]'
    documentTitle = '//*[@placeholder="Enter a title"]'
    uploadFile = '[type="file"]'
    viewDetails = "View Details"
    createAccountButton = "//button[@type='submit' and text()='Create Account']"
    Patient_search_textbox = '//*[@placeholder="Search by Name, Phone, or DOB (mm/dd/yyyy)"]'
    tableRow = '//tbody/tr'
    doctor_search_textbox = '//*[@placeholder="Search by Name"]'
    NextButton = '//button[@type="button" and text()="Next"]'
    CreateOnetimeRXButton = '//*[@data-test="ORDER_FLOW_CREATE_NEW_RX"]'
    createSubscriptionRxButton = '//*[@data-test="ORDER_FLOW_CREATE_CUSTOM_PLAN"]'
    OnetimeSearchMedicine = "New One-time Rx"
    searchForMedication = '//*[@placeholder="Search for a Medication"]'
    AddButton = "//*[@class='is-flex row']//*[text()='Add']"
    addProductButton = "//*[text()='Add']"
    ProductQuantity = "//*[@data-test='amount']//*[@class='Select-arrow']"
    Quantity = '//*[@class="Select-menu-outer"]//div[@aria-label="2"]'
    ProductRefilles = '//*[@data-test="refills"]//*[@class="Select-arrow"]'
    DAWCheckbox = '//*[@class="sc-fzsDOv fUXBxr"]'
    productInstruction = '//*[@placeholder="Enter Rx Instructions"]'
    allergiesButton = '(//*[text()="Edit"])[3]'
    allergiesTextbox = '//*[@class="modal-dialog"]//input'
    doneButton = "//*[text()='Done']"
    addDropchartButton = "//*[@data-test='edit-document']"
    selectDocuments = '(//*[@class="modal-dialog"]//*[@type="checkbox"])[1]'
    selectButton = '//*[@class="modal-dialog"]//*[text()="Select"]'
    skipPayment = '//*[text()="Skip Payment"]'
    providePayment = '//*[text()="Provide Payment"]'
    surgeryDate = '//*[@placeholder="MM/DD/YYYY"]'
    pharmacyNotes = '//*[@placeholder="Enter notes for the pharmacist. (optional) "]'
    submit_CreateOrderButton = '//*[@data-test="ORDER_FLOW_SUBMIT"]'
    submitButton = '//*[text()="Submit"]'
    selectOnetimeTemplate = "(//*[@type='button' and text()='Select'])[1]"
    selectSubscriptionTemplate = "(//*[@type='button' and text()='Select'])[2]"
    templateTitle = '//*[@placeholder="Template Title"]'
    continueButton = "//*[text()='Continue']"
    templateProductSearch = '//*[@placeholder="Search for a Product"]'
    subscriptionProductSearch = '//*[@placeholder="Search for a product"]'
    templateNotes = '//*[@placeholder="Enter notes for the pharmacist"]'
    saveTemplateButton = '//*[text()="Save Template"]'
    accountLinkLabel = 'account-link-label'
    manageUser = 'Manage Users'
    addUserButton = '//*[text()="Add New User"]'
    selectAccountType = 'Select-arrow'
    clerkOption = '//*[@class="Select-menu-outer"]//*[text()="Clerk"]'
    dismissButton = '//*[text()="Dismiss"]'
    mainPatientSearch = 'search'
    mainSearchTextbox = '//*[@placeholder="Search for a Patient by Name or DOB"]'
    mainSearchButton = 'Search-btn'
    cardName = 'name'
    cardNumber = 'number'
    maskCardNumberField = 'masked_card'
    cardCVV = 'cvv'
    addressLine1 = 'street_1'
    addressLine2 = 'street_2'
    addressCity = 'city'
    addressState = '//*[@id="shipping_address"]//*[@class="Select-arrow"]'
    addressZipCode = 'zip'
    state = "//div[@class='Select-menu-outer']//div[@aria-label='" + testData.cell(2, 15).value + "']"
    updateOrderButton = '//*[text()="Update Order"]'

