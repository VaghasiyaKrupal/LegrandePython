from datetime import date
from openpyxl import load_workbook

today = date.today()
currentDate = today.strftime("%m/%d/%Y")

FilePath = "C:/Users/Administrator/PycharmProject/LegrandePython/TestData/Data.xlsx"
datafile = load_workbook(FilePath)
scriptData = datafile['Script Data']


class MasterLocators:
    # Product Creation
    addProductLink = 'Add New Product'
    activeProductCheckbox = '//input[@type="checkbox"]'
    nameField = 'name'
    selectArrow = 'Select-arrow'
    rxProductType = '//div[@aria-label="RX"]'
    otcProductType = '//div[@aria-label="OTC"]'
    compoundProductType = '//div[@aria-label="Compound"]'
    categoryField = 'category'
    brandField = 'brand'
    packageField = 'strength'
    priceField = 'recommended_price'
    gmvField = 'gmv'
    descriptionField = 'description'
    description = "This product is created by the Selenium with Python automation."
    awpField = 'awp_number'
    arpField = 'arp_number'
    ndcField = 'ndc_number'
    pnField = 'pn_number'
    uploadImage = '//input[@type="file"]'
    addProductButton = '//button[text()="Add Product"]'

    # Dispenser Creation
    dispenserScreen = "Dispensers"
    addDispenserLink = 'Add New Dispenser'
    emailField = 'email'
    phoneField = 'phone'
    faxField = 'fax'
    addressStreet1Field = 'address.street_1'
    addressCityField = 'address.city'
    addressZipCodeField = 'address.zip'
    npiField = 'npi_number'
    licenseField = 'reseller_licence_number'
    adminFirstNameField = 'admin.first_name'
    adminLastnameField = 'admin.last_name'
    adminEmailField = 'admin.email'
    adminPhoneField = 'admin.phone'
    selectState = '(//*[@class="Select-arrow"])[1]'
    selectTimezone = '(//*[@class="Select-arrow"])[2]'
    timeZoneValue = '//*[text()="Asia/Kolkata Chennai, Kolkata, Mumbai, New Delhi (GMT+05:30)"]'
    selectHubCategory = '//label[contains(text(),"Hub Dispenser")]'
    dismissButton = '//div[@class="modal-dialog"]//*[text()="Dismiss"]'

    # Practice Creation
    practiceScreen = "Practices"
    addPracticeLink = "Add New Practice"
    dropdownSelectType = '//div[text()="Select..."]'
    practiceValue = '(//*[text()="Practice"])[4]'
    RXDispenser = "//div[@aria-label='"+scriptData.cell(2, 5).value+"']"
    OTCDispenser = "//div[@aria-label='" + scriptData.cell(2, 6).value + "']"
    CompoundDispenser = "//div[@aria-label='" + scriptData.cell(2, 7).value + "']"

    # User Creation
    userScreen = 'Users'
    practiceRadio = '//label[contains(text(),"Practice")]'
    searchPractice = '//*[@placeholder="Search for an account"]'
    selectUserPractice = "(//*[text()='"+scriptData.cell(2, 8).value+"'])[1]"
    userType = "//*[text()='Clerk']"

    # Payor Creation
    PayorScreen = 'Payors & Coupons'
    addPayorButton = '//*[text()="Add New Payor"]'
    insurance = "//*[text()='Insurance']"
    Cash = "//*[text()='Cash']"
    Coupon = "//*[text()='Coupon']"
    payorStatus = "//*[text()='Active']"
    createPayorButton = "//*[text()='Create Payor']"

    # Mailinator Locators
    emailInbox = '//td[contains(text(),"Confirm Legrande Registration")]'
    accountSetupLink = "//a[contains(text(),'Account Setup')]"
    newPassword = 'password'
    confirmPassword = 'password_confirmation'
    setPasswordButton = "//button[contains(text(),'Set Password')]"
    closeButton = "close"
