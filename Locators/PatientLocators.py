from datetime import date
from openpyxl import load_workbook

today = date.today()
currentDate = today.strftime("%m/%d/%Y")

FilePath = "TestData/Data.xlsx"
datafile = load_workbook(FilePath)
testData = datafile['Test Data']


class PatientLocators:
    # Patient Portal Locators
    allergies = 'allergies.0'
    payorAccountNumber = 'account_number'
    payorGroupNumber = 'group_number'
    payorRxBinNumber = 'rx_bin'
    payorPCnNumber = 'pc_number'
    payorPhoneNumber = 'insurer_phone'
    completePayment = '(//*[text()="Complete Payment"])[1]'
    patientAccountLink = '(//*[@fill="currentColor"])[2]'
    patientSettings = '//*[text()="Settings"]'
    patientAccount = "//*[text()='Account']"
    patientPassword = '//*[text()="Password"]'
    patientHealthInsurance = '//*[text()="Health Insurance"]'
    patientAllergies = '//*[text()="Allergies"]'
    patientPreference = '//*[text()="Preferences"]'
    orderScreen = 'orders'
    firstOrder = 'css-5sgxtb'
    saveUpdateButton = "//*[text()='Save Updates']"
    newDate = '(//*[text()="1"])[2]'
    updatePaymentButton = '(//*[text()="Update"])[2]'
    updateShippingButton = '(//*[text()="Update"])[3]'
    CardNumber = 'CardNumber'
    ExperientionDate = 'ExpirationDate'
    cvvCode = "SecurityCode"
    PostalCode = "PostalCode"
    cardDetails = '5555 5537 5304 8194'
    cardExpiryDate = '12/25'
    cardCvvCode = '354'
    cardPostalCode = '584697'
    iframeSaveButton = '//*[text()="Save"]'
    shippingState = '//input[@id="state"]'
    stateValue = "//*[text()='" + testData.cell(2, 20).value + "']"
    currentPasswordID = 'old_password'
    newPasswordID = 'new_password'
    confirmPasswordID = 'confirm_password'
    addNewInsurance = '//*[text()="Add New Insurance"]'
    payorID = 'payor_id'
    firstPayor = '//form[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
    editPayorName = '//form[1]/div[1]/div[1]/div[1]/div[2]/div[2]'
    addInsuranceButton = '//*[@type="submit" and text()="Add Insurance"]'
    allergiesID = 'allergy'
    addAllergyButton = '//*[text()="Add Allergy"]'
    removeAllergyButton = '//*[@aria-label="remove-allergy"]'
    confirmButton = '//*[text()="Confirm"]'
    messageCheckbox = '//*[@id="__next"]/div/div[2]/div[1]/div[2]/div[2]/label/span'
    additionalCheckbox = '//*[@id="__next"]/div/div[2]/div[1]/div[3]/div[2]/label/span/span'
    cancelPlanCheckbox = '//*[text()="Permanently cancel my plan."]'
    reasoneDropdownID = 'selected_remarks'
    selectOtherReason = '//*[text()="Other"]'
    userReasonID = 'user_remarks'
    myQueueOrderDate = "(//*[@class='table-row']//td[2][contains(text(),'" + currentDate + "')])[1]"
    myQueuePatientName = "(//*[@class='table-row']//td[3][contains(text(),'" + testData.cell(2, 1).value + "')])[1]"
